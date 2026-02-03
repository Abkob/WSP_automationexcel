"""
Utility functions for data processing and file handling.
Supports: Excel (.xlsx, .xls, .xlsm), CSV, TSV, JSON
"""

import os
import datetime
import numpy as np
import pandas as pd
from typing import Optional, Tuple
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill


DATE_COL_NAME = "Date Added"


def smart_cast_dtypes(df: pd.DataFrame) -> pd.DataFrame:
    """Intelligently detect and cast column data types."""
    df = df.copy()
    df = df.convert_dtypes()
    
    for col in df.columns:
        series = df[col]
        
        if pd.api.types.is_numeric_dtype(series.dtype):
            continue
        if pd.api.types.is_datetime64_any_dtype(series.dtype):
            continue
        
        if series.dtype == "string" or series.dtype == object:
            series_str = series.astype(str)
            
            has_date_chars = series_str.str.contains(r"[/-]", regex=True, na=False).mean()
            
            if has_date_chars >= 0.5:
                try:
                    dt_series = pd.to_datetime(series_str, errors="coerce")
                    non_na_ratio = dt_series.notna().mean()
                    
                    if non_na_ratio >= 0.5 and dt_series.notna().sum() > 0:
                        df[col] = dt_series
                        continue
                except Exception:
                    pass
            
            try:
                num_series = pd.to_numeric(series_str.replace("", np.nan), errors="coerce")
                non_na_ratio = num_series.notna().mean()
                
                if non_na_ratio >= 0.5 and num_series.notna().sum() > 0:
                    df[col] = num_series
            except Exception:
                pass
    
    return df


def load_dataframe_from_file(filepath: str, sheet_name: Optional[str] = None) -> Tuple[pd.DataFrame, bool, list]:
    """Load DataFrame from Excel, CSV, TSV, or JSON file."""
    filepath_lower = filepath.lower()
    
    try:
        # Excel files
        if filepath_lower.endswith(('.xlsx', '.xls', '.xlsm')):
            excel_file = pd.ExcelFile(filepath)
            sheet_names = excel_file.sheet_names
            
            if sheet_name is None and len(sheet_names) > 1:
                return pd.DataFrame(), True, sheet_names
            
            sheet = sheet_name if sheet_name else sheet_names[0]
            df = pd.read_excel(excel_file, sheet_name=sheet)
            
        # CSV files
        elif filepath_lower.endswith('.csv'):
            try:
                df = pd.read_csv(filepath)
            except UnicodeDecodeError:
                df = pd.read_csv(filepath, encoding='latin1')
            sheet_names = []
        
        # TSV files (NEW)
        elif filepath_lower.endswith('.tsv'):
            try:
                df = pd.read_csv(filepath, sep='\t')
            except UnicodeDecodeError:
                df = pd.read_csv(filepath, sep='\t', encoding='latin1')
            sheet_names = []
        
        # JSON files (NEW)
        elif filepath_lower.endswith('.json'):
            df = pd.read_json(filepath)
            sheet_names = []
        
        else:
            raise ValueError(f"Unsupported file format: {filepath}")
        
        df.columns = [str(c) for c in df.columns]
        df = smart_cast_dtypes(df)
        
        return df, False, sheet_names
        
    except Exception as e:
        raise Exception(f"Failed to load file: {e}")


def add_date_column(df: pd.DataFrame) -> pd.DataFrame:
    """Add or update the Date Added column."""
    df = df.copy()
    now = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    df[DATE_COL_NAME] = now
    return df


def merge_dataframes(existing_df: pd.DataFrame, new_df: pd.DataFrame) -> Tuple[pd.DataFrame, bool, str]:
    """Merge new DataFrame with existing one."""
    existing_base = [c for c in existing_df.columns if c != DATE_COL_NAME]
    new_base = [c for c in new_df.columns if c != DATE_COL_NAME]
    
    if set(existing_base) != set(new_base):
        missing_in_new = set(existing_base) - set(new_base)
        extra_in_new = set(new_base) - set(existing_base)
        
        msg = "Column mismatch:\n"
        if missing_in_new:
            msg += f"Missing: {', '.join(missing_in_new)}\n"
        if extra_in_new:
            msg += f"Extra: {', '.join(extra_in_new)}\n"
        
        return existing_df, False, msg
    
    new_df = new_df[existing_base + [DATE_COL_NAME]]
    merged = pd.concat([existing_df, new_df], ignore_index=True)
    merged = merged.drop_duplicates(subset=existing_base, keep="last")
    merged = merged.reset_index(drop=True)
    
    return merged, True, "Successfully merged"


def export_to_excel_formatted(df: pd.DataFrame, filepath: str, filter_manager,
                              split_sheets: bool = True, filter_mode: str = "all",
                              mask_override: Optional[np.ndarray] = None):
    """Export DataFrame to Excel with formatting."""
    if df.empty:
        with pd.ExcelWriter(filepath, engine="openpyxl") as writer:
            pd.DataFrame().to_excel(writer, index=False, sheet_name="ALL")
        return
    
    if mask_override is not None:
        try:
            mask = np.asarray(mask_override, dtype=bool)
        except Exception:
            mask = np.array([False] * len(df), dtype=bool)
    elif filter_manager is None or not getattr(filter_manager, "has_filters", lambda: False)():
        mask = np.array([False] * len(df), dtype=bool)
    else:
        if filter_mode == "any":
            matcher = filter_manager.matches_any_filter
        elif hasattr(filter_manager, "matches_all_filters"):
            matcher = filter_manager.matches_all_filters
        else:
            matcher = filter_manager.matches_any_filter

        mask = np.array([
            matcher(df.iloc[i])
            for i in range(len(df))
        ], dtype=bool)
    
    if split_sheets:
        sheets = {
            "ALL": df.copy(),
            "HIGHLIGHTED": df[mask].copy(),
            "UNHIGHLIGHTED": df[~mask].copy()
        }
    else:
        sheets = {"Data": df.copy()}
    
    with pd.ExcelWriter(filepath, engine="openpyxl") as writer:
        for sheet_name, sheet_df in sheets.items():
            sheet_df.to_excel(writer, index=False, sheet_name=sheet_name)
            ws = writer.sheets[sheet_name]
            
            ws.freeze_panes = "A2"
            
            header_font = Font(bold=True)
            for cell in ws[1]:
                cell.font = header_font
            
            for col_idx, column_cells in enumerate(ws.columns, start=1):
                max_length = 0
                for cell in column_cells:
                    if cell.value is not None:
                        max_length = max(max_length, len(str(cell.value)))
                
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[get_column_letter(col_idx)].width = adjusted_width
            
            _apply_filter_highlights(ws, sheet_df, filter_manager)
            ws.auto_filter.ref = ws.dimensions


def _apply_filter_highlights(worksheet, df: pd.DataFrame, filter_manager):
    """Apply conditional formatting based on active filters."""
    if df.empty:
        return
    
    col_name_to_idx = {name: idx + 1 for idx, name in enumerate(df.columns)}
    all_filters = filter_manager.get_all_filters()
    
    for filter_rule in all_filters:
        col_name = filter_rule.column
        if col_name not in col_name_to_idx:
            continue
        
        col_idx = col_name_to_idx[col_name]
        color = filter_rule.get_color()
        rgb_hex = f"{color.red():02X}{color.green():02X}{color.blue():02X}"
        fill = PatternFill(start_color=rgb_hex, end_color=rgb_hex, fill_type="solid")
        
        for row_idx in range(2, worksheet.max_row + 1):
            cell = worksheet.cell(row=row_idx, column=col_idx)
            
            if cell.value is not None and filter_rule.matches(cell.value):
                if cell.fill is None or cell.fill.fill_type is None:
                    cell.fill = fill


def create_archive_snapshot(df: pd.DataFrame, filter_manager, archive_dir: str = "archives"):
    """Create a timestamped archive snapshot."""
    os.makedirs(archive_dir, exist_ok=True)
    
    timestamp = datetime.datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
    filename = f"archive_{timestamp}.xlsx"
    filepath = os.path.join(archive_dir, filename)
    
    export_to_excel_formatted(df, filepath, filter_manager, split_sheets=True)
    
    return filepath


def get_archive_list(archive_dir: str = "archives") -> list:
    """Get list of archive files with metadata."""
    if not os.path.exists(archive_dir):
        return []
    
    archives = []
    
    for filename in os.listdir(archive_dir):
        if filename.startswith("archive_") and filename.endswith(".xlsx"):
            filepath = os.path.join(archive_dir, filename)
            
            try:
                timestamp_str = filename.replace("archive_", "").replace(".xlsx", "")
                timestamp = datetime.datetime.strptime(timestamp_str, "%Y-%m-%d_%H-%M-%S")
            except Exception:
                timestamp = None
            
            try:
                size = os.path.getsize(filepath)
            except Exception:
                size = 0
            
            archives.append({
                'filename': filename,
                'path': filepath,
                'timestamp': timestamp,
                'size': size
            })
    
    archives.sort(key=lambda x: x['timestamp'] if x['timestamp'] else datetime.datetime.min, reverse=True)
    
    return archives


def format_file_size(size_bytes: int) -> str:
    """Format file size in human-readable format."""
    for unit in ['B', 'KB', 'MB', 'GB']:
        if size_bytes < 1024.0:
            return f"{size_bytes:.1f} {unit}"
        size_bytes /= 1024.0
    return f"{size_bytes:.1f} TB"


def save_filters_to_file(filter_manager, filepath: str = "app_filters.json"):
    """Save current filters to a JSON file."""
    import json
    
    try:
        all_filters = filter_manager.get_all_filters()
        filters_data = [f.to_dict() for f in all_filters]
        
        with open(filepath, 'w') as f:
            json.dump(filters_data, f, indent=2)
        
        return True
    except Exception as e:
        print(f"Error saving filters: {e}")
        return False


def load_filters_from_file(filepath: str = "app_filters.json"):
    """Load filters from JSON file."""
    import json
    from models import NumericFilter, TextFilter, DateFilter
    
    if not os.path.exists(filepath):
        return []
    
    try:
        with open(filepath, 'r') as f:
            filters_data = json.load(f)
        
        filters = []
        for filter_data in filters_data:
            filter_type = filter_data.get("type")
            
            if filter_type == "numeric":
                filters.append(NumericFilter.from_dict(filter_data))
            elif filter_type == "text":
                filters.append(TextFilter.from_dict(filter_data))
            elif filter_type == "date":
                filters.append(DateFilter.from_dict(filter_data))
        
        return filters
    except Exception as e:
        print(f"Error loading filters: {e}")
        return []


def get_last_loaded_file(filepath: str = "app_last_file.txt") -> Optional[str]:
    """Get the path of the last loaded file."""
    if not os.path.exists(filepath):
        return None
    
    try:
        with open(filepath, 'r') as f:
            last_file = f.read().strip()
        
        if os.path.exists(last_file):
            return last_file
        return None
    except Exception:
        return None


def save_last_loaded_file(file_path: str, filepath: str = "app_last_file.txt"):
    """Save the path of the last loaded file."""
    try:
        with open(filepath, 'w') as f:
            f.write(file_path)
        return True
    except Exception as e:
        print(f"Error saving last file: {e}")
        return False
