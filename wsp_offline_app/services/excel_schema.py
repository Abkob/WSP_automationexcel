from __future__ import annotations

import re
from dataclasses import dataclass
from pathlib import Path
from typing import Iterable

from openpyxl import load_workbook


EXPECTED_WSP_COLUMNS: tuple[str, ...] = (
    "STUD_ID",
    "STUD_NAME",
    "MAJR_DESC",
    "CLAS_DESC",
    "STUD_EMAIL",
    "WSP_WRITTEN_LANGUAGES",
    "WSP_SPOKEN_LANGUAGES",
    "WSP_ORGANIZATIONAL_SKILLS",
    "WSP_TECHNICAL_SKILLS",
    "WSP_INTERPERSONAL_SKILLS",
    "WSP_ADDITIONAL_SKILLS",
    "WSP_PREV_WORK",
    "WSP_PREVIOUS_TYPE_OF_WORK",
    "WSP_PREFERRED_TYPE_OF_WORK",
    "DEANS_WARNING",
    "ENRL_TERM",
    "DEAN_WARN",
    "MOBILE_NBR",
    "PROBATION",
    "APPLICATION_DATE",
    "CUM_GPA",
    "STST_DESC",
    "STYP_CODE",
    "STYP_DESC",
    "ENROLLED_IND",
    "REGISTERED_IND",
    "LEVL_CODE",
    "COLL_CODE",
    "TOTAL_CREDIT_HOURS",
    "ASTD_TERM",
    "ATSD_CODE_END_OF_TERM",
    "ASTD_DESC",
    "ASTD_DATE_END_OF_TERM",
    "USAID",
    "MASTER_CARD",
    "UPP_MEPI",
    "GAS",
    "FINANCIAL_AID",
    "DORMS",
)


SUPPORTED_EXCEL_EXTENSIONS = {".xlsx", ".xlsm"}


@dataclass(frozen=True)
class WorkbookSchema:
    path: Path
    sheet_names: tuple[str, ...]
    active_sheet: str
    row_count: int
    data_row_count: int
    column_count: int
    headers: tuple[str, ...]


@dataclass(frozen=True)
class SchemaComparison:
    expected_columns: tuple[str, ...]
    actual_columns: tuple[str, ...]
    missing_columns: tuple[str, ...]
    new_columns: tuple[str, ...]
    duplicate_columns: tuple[str, ...]

    @property
    def matches_expected(self) -> bool:
        return not self.missing_columns and not self.new_columns and not self.duplicate_columns


@dataclass(frozen=True)
class HeaderCleaningResult:
    original_headers: tuple[str, ...]
    normalized_headers: tuple[str, ...]
    original_by_normalized: dict[str, str]
    duplicate_headers: tuple[str, ...]


class DuplicateHeaderError(ValueError):
    """Raised when two or more raw headers normalize to the same database-safe name."""


def normalize_header(value: object) -> str:
    text = "" if value is None else str(value)
    text = text.strip().upper()
    text = re.sub(r"[^A-Z0-9]+", "_", text)
    text = re.sub(r"_+", "_", text)
    return text.strip("_")


def normalize_headers(values: Iterable[object]) -> tuple[str, ...]:
    return tuple(normalize_header(value) for value in values)


def find_duplicates(values: Iterable[str]) -> tuple[str, ...]:
    seen: set[str] = set()
    duplicates: list[str] = []

    for value in values:
        if value in seen and value not in duplicates:
            duplicates.append(value)
        seen.add(value)

    return tuple(duplicates)


def clean_headers(raw_headers: Iterable[object], *, reject_duplicates: bool = True) -> HeaderCleaningResult:
    original_headers = tuple("" if value is None else str(value).strip() for value in raw_headers)
    normalized_headers = normalize_headers(original_headers)
    duplicate_headers = find_duplicates(header for header in normalized_headers if header)

    if duplicate_headers and reject_duplicates:
        duplicates = ", ".join(duplicate_headers)
        raise DuplicateHeaderError(f"Duplicate headers after cleaning: {duplicates}")

    original_by_normalized = {
        normalized: original
        for original, normalized in zip(original_headers, normalized_headers, strict=True)
        if normalized
    }

    return HeaderCleaningResult(
        original_headers=original_headers,
        normalized_headers=normalized_headers,
        original_by_normalized=original_by_normalized,
        duplicate_headers=duplicate_headers,
    )


def compare_headers(
    actual_headers: Iterable[object],
    expected_headers: Iterable[str] = EXPECTED_WSP_COLUMNS,
) -> SchemaComparison:
    actual = normalize_headers(actual_headers)
    expected = tuple(expected_headers)
    actual_set = set(actual)

    return SchemaComparison(
        expected_columns=expected,
        actual_columns=actual,
        missing_columns=tuple(column for column in expected if column not in actual_set),
        new_columns=tuple(column for column in actual if column and column not in expected),
        duplicate_columns=find_duplicates(column for column in actual if column),
    )


def read_workbook_schema(path: Path | str, sheet_name: str | None = None) -> WorkbookSchema:
    workbook_path = Path(path)

    if not workbook_path.exists():
        raise FileNotFoundError(f"Excel workbook does not exist: {workbook_path}")

    if workbook_path.suffix.lower() not in SUPPORTED_EXCEL_EXTENSIONS:
        raise ValueError(f"Unsupported Excel extension: {workbook_path.suffix}")

    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    active_sheet_name = sheet_name or workbook.sheetnames[0]

    if active_sheet_name not in workbook.sheetnames:
        raise ValueError(f"Sheet {active_sheet_name!r} was not found in {workbook_path}")

    worksheet = workbook[active_sheet_name]
    raw_headers, row_count = read_header_row_and_row_count(worksheet)
    headers = normalize_headers(raw_headers)

    return WorkbookSchema(
        path=workbook_path,
        sheet_names=tuple(workbook.sheetnames),
        active_sheet=active_sheet_name,
        row_count=row_count,
        data_row_count=max(row_count - 1, 0),
        column_count=len(headers),
        headers=headers,
    )


def read_header_row_and_row_count(worksheet) -> tuple[tuple[object, ...], int]:
    if worksheet.max_column is not None and worksheet.max_row is not None:
        headers = tuple(
            worksheet.cell(row=1, column=index).value
            for index in range(1, worksheet.max_column + 1)
        )
        return headers, worksheet.max_row

    rows = list(worksheet.iter_rows(values_only=True))
    if not rows:
        return (), 0
    return tuple(rows[0]), len(rows)
