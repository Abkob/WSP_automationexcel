from __future__ import annotations

import time
import json
from hashlib import sha256
from dataclasses import dataclass
from datetime import UTC, datetime
from pathlib import Path
from typing import Callable, Mapping, TypeVar

import pandas as pd
from sqlalchemy.orm import Session

from database.models import FileImportLog, ImportBatch, StudentCurrent, StudentHistory
from services.excel_schema import EXPECTED_WSP_COLUMNS, SUPPORTED_EXCEL_EXTENSIONS
from services.value_normalizer import is_empty_value, normalize_text


ROW_HASH_EXCLUDED_FIELDS = {
    "created_at",
    "updated_at",
    "row_hash",
    "first_seen_batch_id",
    "last_seen_batch_id",
    "missing_from_latest_import",
    "added_to_db_at",
    "modified_in_db_at",
}

CURRENT_STUDENT_COLUMNS = set(EXPECTED_WSP_COLUMNS)


class ExcelIntakeError(Exception):
    """Base error for Excel intake failures."""


class TemporaryExcelFileIgnored(ExcelIntakeError):
    """Raised when an Excel temporary lock file should be ignored."""


class UnsupportedExcelFileError(ExcelIntakeError):
    """Raised when the supplied file path is not a supported Excel workbook."""


class FileNotStableError(ExcelIntakeError):
    """Raised when a file keeps changing and should not be imported yet."""


class DuplicateExcelFileError(ExcelIntakeError):
    """Raised when the exact same file was already imported."""


class GeneratedExportWorkbookError(ExcelIntakeError):
    """Raised when a WSP filtered export is submitted as a source workbook."""


class MissingStudentIdError(ExcelIntakeError):
    """Raised when an imported row does not contain a usable STUD_ID."""


class NoValidStudentRowsError(ExcelIntakeError):
    """Raised when a workbook has no usable student records to import."""


def is_generated_export_filename(path: str | Path) -> bool:
    """Recognize filenames produced by WSP's own export workflow."""
    stem = Path(path).stem.casefold()
    return "filtered_students" in stem or "filtered_results" in stem


def reject_generated_export_workbook(path: str | Path) -> None:
    """Prevent an exported subset from replacing the master student workbook."""
    workbook_path = Path(path)
    if is_generated_export_filename(workbook_path):
        raise GeneratedExportWorkbookError(
            f"Import rejected: {workbook_path.name} is a WSP filtered export, not a master student workbook. "
            "Use the original full WSP workbook in the Import Folder."
        )

    try:
        from openpyxl import load_workbook

        workbook = load_workbook(workbook_path, read_only=True, data_only=True)
        try:
            sheet_names = set(workbook.sheetnames)
        finally:
            workbook.close()
    except GeneratedExportWorkbookError:
        raise
    except Exception:
        return

    if {"Filtered Results", "Filter Metadata"}.issubset(sheet_names):
        raise GeneratedExportWorkbookError(
            f"Import rejected: {workbook_path.name} contains WSP Filtered Results and Filter Metadata sheets. "
            "Exports cannot replace the master student workbook."
        )


@dataclass(frozen=True)
class ExcelIntakeResult:
    path: Path
    file_size_bytes: int
    extension: str


@dataclass(frozen=True)
class ExcelWorkbookData:
    path: Path
    sheet_names: tuple[str, ...]
    active_sheet: str
    headers: tuple[str, ...]
    rows: tuple[dict[str, object | None], ...]

    @property
    def row_count(self) -> int:
        return len(self.rows)

    @property
    def column_count(self) -> int:
        return len(self.headers)


@dataclass(frozen=True)
class StudentUpsertResult:
    stud_id: str
    change_type: str
    row_hash: str


@dataclass(frozen=True)
class MissingStudentDetectionResult:
    newly_missing_student_ids: tuple[str, ...]
    already_missing_student_ids: tuple[str, ...]


SizeReader = Callable[[Path], int]
SleepFunc = Callable[[float], None]
ClockFunc = Callable[[], float]
ImportOperationResult = TypeVar("ImportOperationResult")
ImportOperation = Callable[[Session, ImportBatch], ImportOperationResult]


def is_temporary_excel_file(path: Path | str) -> bool:
    return Path(path).name.startswith("~$")


def validate_excel_file_path(path: Path | str) -> Path:
    workbook_path = Path(path).resolve()

    if is_temporary_excel_file(workbook_path):
        raise TemporaryExcelFileIgnored(f"Ignoring temporary Excel file: {workbook_path}")

    if not workbook_path.exists():
        raise FileNotFoundError(f"Excel file does not exist: {workbook_path}")

    if not workbook_path.is_file():
        raise UnsupportedExcelFileError(f"Excel path is not a file: {workbook_path}")

    if workbook_path.suffix.lower() not in SUPPORTED_EXCEL_EXTENSIONS:
        raise UnsupportedExcelFileError(f"Unsupported Excel extension: {workbook_path.suffix}")

    return workbook_path


def _default_size_reader(path: Path) -> int:
    return path.stat().st_size


def wait_for_file_size_to_stabilize(
    path: Path | str,
    *,
    required_stable_reads: int = 2,
    poll_interval_seconds: float = 0.25,
    timeout_seconds: float = 10.0,
    size_reader: SizeReader = _default_size_reader,
    sleep_func: SleepFunc = time.sleep,
    clock_func: ClockFunc = time.monotonic,
) -> int:
    if required_stable_reads < 2:
        raise ValueError("required_stable_reads must be at least 2")
    if poll_interval_seconds < 0:
        raise ValueError("poll_interval_seconds cannot be negative")
    if timeout_seconds < 0:
        raise ValueError("timeout_seconds cannot be negative")

    workbook_path = Path(path).resolve()
    start = clock_func()
    last_size: int | None = None
    stable_reads = 0

    while clock_func() - start <= timeout_seconds:
        current_size = size_reader(workbook_path)

        if current_size == last_size:
            stable_reads += 1
        else:
            last_size = current_size
            stable_reads = 1

        if stable_reads >= required_stable_reads:
            return current_size

        sleep_func(poll_interval_seconds)

    raise FileNotStableError(f"Excel file size did not stabilize before timeout: {workbook_path}")


def intake_excel_file(
    path: Path | str,
    *,
    required_stable_reads: int = 2,
    poll_interval_seconds: float = 0.25,
    timeout_seconds: float = 10.0,
) -> ExcelIntakeResult:
    workbook_path = validate_excel_file_path(path)
    file_size = wait_for_file_size_to_stabilize(
        workbook_path,
        required_stable_reads=required_stable_reads,
        poll_interval_seconds=poll_interval_seconds,
        timeout_seconds=timeout_seconds,
    )

    return ExcelIntakeResult(
        path=workbook_path,
        file_size_bytes=file_size,
        extension=workbook_path.suffix.lower(),
    )


def calculate_file_hash(path: Path | str, *, chunk_size: int = 1024 * 1024) -> str:
    if chunk_size <= 0:
        raise ValueError("chunk_size must be positive")

    workbook_path = Path(path).resolve()
    digest = sha256()

    with workbook_path.open("rb") as handle:
        while chunk := handle.read(chunk_size):
            digest.update(chunk)

    return digest.hexdigest()


def find_import_batch_by_hash(session: Session, file_hash: str) -> ImportBatch | None:
    return session.query(ImportBatch).filter_by(file_hash=file_hash).one_or_none()


def log_import_event(
    session: Session,
    *,
    batch_id: int | None,
    event_type: str,
    message: str,
    row_number: int | None = None,
    stud_id: str | None = None,
    details_json: dict | None = None,
) -> FileImportLog:
    event = FileImportLog(
        batch_id=batch_id,
        event_type=event_type,
        row_number=row_number,
        STUD_ID=stud_id,
        message=message,
        details_json=details_json or {},
    )
    session.add(event)
    return event


def ensure_file_not_previously_imported(
    session: Session,
    path: Path | str,
    *,
    file_hash: str | None = None,
    log_duplicate: bool = True,
) -> str:
    workbook_path = Path(path).resolve()
    resolved_hash = file_hash or calculate_file_hash(workbook_path)
    existing_batch = find_import_batch_by_hash(session, resolved_hash)

    if existing_batch is not None:
        if log_duplicate:
            log_import_event(
                session,
                batch_id=existing_batch.batch_id,
                event_type="duplicate_file_skipped",
                message=f"Skipped duplicate Excel file: {workbook_path}",
                details_json={
                    "file_hash": resolved_hash,
                    "original_batch_id": existing_batch.batch_id,
                    "duplicate_path": str(workbook_path),
                },
            )
        raise DuplicateExcelFileError(
            f"Excel file was already imported in batch {existing_batch.batch_id}: {workbook_path}"
        )

    return resolved_hash


def _none_if_missing(value: object) -> object | None:
    if pd.isna(value):
        return None
    return value


def read_excel_workbook(path: Path | str, sheet_name: str | None = None) -> ExcelWorkbookData:
    workbook_path = validate_excel_file_path(path)
    with pd.ExcelFile(workbook_path, engine="openpyxl") as excel_file:
        active_sheet = sheet_name or excel_file.sheet_names[0]

        if active_sheet not in excel_file.sheet_names:
            raise ValueError(f"Sheet {active_sheet!r} was not found in {workbook_path}")

        frame = pd.read_excel(
            excel_file,
            sheet_name=active_sheet,
            dtype=object,
            engine="openpyxl",
        )
        headers = tuple(str(column) for column in frame.columns)
        rows = tuple(
            {
                header: _none_if_missing(value)
                for header, value in zip(headers, row, strict=True)
            }
            for row in frame.itertuples(index=False, name=None)
        )
        sheet_names = tuple(excel_file.sheet_names)

    return ExcelWorkbookData(
        path=workbook_path,
        sheet_names=sheet_names,
        active_sheet=active_sheet,
        headers=headers,
        rows=rows,
    )


def normalize_student_id(value: object) -> str | None:
    if is_empty_value(value) or isinstance(value, bool):
        return None

    if isinstance(value, float) and value.is_integer():
        return str(int(value))

    text = normalize_text(value)
    if text is None:
        return None

    if text.endswith(".0") and text[:-2].isdigit():
        return text[:-2]

    return text


def extract_required_student_id(
    row: Mapping[str, object | None],
    *,
    session: Session | None = None,
    batch_id: int | None = None,
    row_number: int | None = None,
) -> str:
    student_id = normalize_student_id(row.get("STUD_ID"))

    if student_id:
        return student_id

    if session is not None:
        log_import_event(
            session,
            batch_id=batch_id,
            event_type="row_rejected",
            row_number=row_number,
            message="Missing or invalid STUD_ID",
            details_json={
                "reason": "missing_stud_id",
                "row": dict(row),
            },
        )

    raise MissingStudentIdError(f"Row {row_number or '?'} does not contain a usable STUD_ID")


def build_row_hash_payload(
    row: Mapping[str, object | None],
    *,
    excluded_fields: set[str] | None = None,
) -> dict[str, object | None]:
    exclusions = excluded_fields or ROW_HASH_EXCLUDED_FIELDS
    return {
        key: row[key]
        for key in sorted(row)
        if key not in exclusions
    }


def generate_row_hash(row: Mapping[str, object | None]) -> str:
    payload = build_row_hash_payload(row)
    serialized = json.dumps(payload, sort_keys=True, separators=(",", ":"), default=str)
    return sha256(serialized.encode("utf-8")).hexdigest()


def split_current_and_extra_columns(row: Mapping[str, object | None]) -> tuple[dict[str, object | None], dict[str, object | None]]:
    current_values = {
        key: value
        for key, value in row.items()
        if key in CURRENT_STUDENT_COLUMNS
    }
    extra_values = {
        key: value
        for key, value in row.items()
        if key not in CURRENT_STUDENT_COLUMNS and key not in ROW_HASH_EXCLUDED_FIELDS
    }
    return current_values, extra_values


def snapshot_student_current(student: StudentCurrent) -> dict[str, object | None]:
    snapshot: dict[str, object | None] = {
        column: getattr(student, column)
        for column in EXPECTED_WSP_COLUMNS
    }
    snapshot["extra_columns_json"] = student.extra_columns_json
    snapshot["row_hash"] = student.row_hash
    snapshot["first_seen_batch_id"] = student.first_seen_batch_id
    snapshot["last_seen_batch_id"] = student.last_seen_batch_id
    snapshot["missing_from_latest_import"] = student.missing_from_latest_import
    snapshot["added_to_db_at"] = student.added_to_db_at.isoformat() if student.added_to_db_at else None
    snapshot["modified_in_db_at"] = student.modified_in_db_at.isoformat() if student.modified_in_db_at else None
    return snapshot


def apply_row_to_student(
    student: StudentCurrent,
    row: Mapping[str, object | None],
    row_hash: str,
    batch_id: int | None,
    *,
    modified_at: datetime | None = None,
) -> None:
    current_values, extra_values = split_current_and_extra_columns(row)
    for column_name, value in current_values.items():
        setattr(student, column_name, value)

    student.extra_columns_json = extra_values
    student.row_hash = row_hash
    student.last_seen_batch_id = batch_id
    student.missing_from_latest_import = False
    if modified_at is not None:
        student.modified_in_db_at = modified_at


def upsert_student_row(session: Session, row: Mapping[str, object | None], *, batch_id: int | None) -> StudentUpsertResult:
    stud_id = extract_required_student_id(row)
    row_with_id = dict(row)
    row_with_id["STUD_ID"] = stud_id
    row_hash = generate_row_hash(row_with_id)

    student = session.query(StudentCurrent).filter_by(STUD_ID=stud_id).one_or_none()

    if student is None:
        added_at = datetime.now(UTC)
        current_values, extra_values = split_current_and_extra_columns(row_with_id)
        student = StudentCurrent(
            **current_values,
            extra_columns_json=extra_values,
            row_hash=row_hash,
            first_seen_batch_id=batch_id,
            last_seen_batch_id=batch_id,
            missing_from_latest_import=False,
            added_to_db_at=added_at,
            modified_in_db_at=None,
        )
        session.add(student)
        return StudentUpsertResult(stud_id=stud_id, change_type="new_student", row_hash=row_hash)

    was_missing = student.missing_from_latest_import
    if student.row_hash == row_hash:
        student.last_seen_batch_id = batch_id
        student.missing_from_latest_import = False
        return StudentUpsertResult(
            stud_id=stud_id,
            change_type="restored_student" if was_missing else "unchanged_student",
            row_hash=row_hash,
        )

    modified_at = datetime.now(UTC)
    history = StudentHistory(
        batch_id=batch_id,
        STUD_ID=stud_id,
        all_excel_columns=snapshot_student_current(student),
        row_hash=student.row_hash,
        valid_from=student.created_at,
        valid_to=modified_at,
        is_current=False,
        change_type="updated_student",
    )
    session.add(history)
    apply_row_to_student(student, row_with_id, row_hash, batch_id, modified_at=modified_at)

    return StudentUpsertResult(
        stud_id=stud_id,
        change_type="updated_student",
        row_hash=row_hash,
    )


def mark_missing_students(
    session: Session,
    *,
    seen_student_ids: set[str],
    batch_id: int | None,
) -> MissingStudentDetectionResult:
    newly_missing: list[str] = []
    already_missing: list[str] = []

    for student in session.query(StudentCurrent).all():
        if student.STUD_ID in seen_student_ids:
            continue

        if student.missing_from_latest_import:
            already_missing.append(student.STUD_ID)
            continue

        session.add(
            StudentHistory(
                batch_id=batch_id,
                STUD_ID=student.STUD_ID,
                all_excel_columns=snapshot_student_current(student),
                row_hash=student.row_hash,
                valid_from=student.created_at,
                valid_to=datetime.now(UTC),
                is_current=False,
                change_type="missing_from_latest_import",
            )
        )
        student.missing_from_latest_import = True
        student.last_seen_batch_id = batch_id
        newly_missing.append(student.STUD_ID)

    return MissingStudentDetectionResult(
        newly_missing_student_ids=tuple(newly_missing),
        already_missing_student_ids=tuple(already_missing),
    )


def execute_import_transaction(
    session_factory: Callable[[], Session],
    *,
    batch_id: int,
    operation: ImportOperation[ImportOperationResult],
) -> ImportOperationResult:
    try:
        with session_factory() as session:
            with session.begin():
                batch = session.get(ImportBatch, batch_id)
                if batch is None:
                    raise ValueError(f"Import batch does not exist: {batch_id}")
                batch.status = "running"
                result = operation(session, batch)
                batch.status = "completed"
                return result
    except Exception as exc:
        with session_factory() as failure_session:
            batch = failure_session.get(ImportBatch, batch_id)
            if batch is not None:
                batch.status = "failed"
                batch.error_message = str(exc)
                failure_session.commit()
        raise
