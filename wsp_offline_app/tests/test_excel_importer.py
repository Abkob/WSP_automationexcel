from __future__ import annotations

from collections import deque
from pathlib import Path

import pytest
from openpyxl import Workbook

from database.db import create_session_factory, create_sqlite_engine, initialize_database
from database.models import FileImportLog, ImportBatch, StudentCurrent
from database.models import StudentHistory
from services.excel_importer import (
    DuplicateExcelFileError,
    FileNotStableError,
    GeneratedExportWorkbookError,
    MissingStudentIdError,
    TemporaryExcelFileIgnored,
    UnsupportedExcelFileError,
    calculate_file_hash,
    ensure_file_not_previously_imported,
    execute_import_transaction,
    extract_required_student_id,
    find_import_batch_by_hash,
    generate_row_hash,
    intake_excel_file,
    is_generated_export_filename,
    is_temporary_excel_file,
    mark_missing_students,
    normalize_student_id,
    read_excel_workbook,
    reject_generated_export_workbook,
    upsert_student_row,
    validate_excel_file_path,
    wait_for_file_size_to_stabilize,
)


def create_minimal_workbook(path: Path) -> Path:
    workbook = Workbook()
    workbook.active.append(["STUD_ID"])
    workbook.save(path)
    return path


def test_valid_xlsx_file_is_accepted(tmp_path: Path) -> None:
    path = create_minimal_workbook(tmp_path / "WSP.xlsx")

    result = intake_excel_file(path, poll_interval_seconds=0.001, timeout_seconds=1)

    assert result.path == path.resolve()
    assert result.extension == ".xlsx"
    assert result.file_size_bytes > 0


def test_missing_file_raises_clear_error(tmp_path: Path) -> None:
    with pytest.raises(FileNotFoundError, match="Excel file does not exist"):
        validate_excel_file_path(tmp_path / "missing.xlsx")


def test_txt_file_is_rejected(tmp_path: Path) -> None:
    path = tmp_path / "not_excel.txt"
    path.write_text("hello", encoding="utf-8")

    with pytest.raises(UnsupportedExcelFileError, match="Unsupported Excel extension"):
        validate_excel_file_path(path)


def test_directory_path_is_rejected(tmp_path: Path) -> None:
    with pytest.raises(UnsupportedExcelFileError, match="not a file"):
        validate_excel_file_path(tmp_path)


def test_temporary_excel_file_is_ignored(tmp_path: Path) -> None:
    path = create_minimal_workbook(tmp_path / "~$WSP.xlsx")

    assert is_temporary_excel_file(path)
    with pytest.raises(TemporaryExcelFileIgnored, match="Ignoring temporary Excel file"):
        validate_excel_file_path(path)


def test_generated_filtered_export_is_rejected_by_filename(tmp_path: Path) -> None:
    path = create_minimal_workbook(tmp_path / "20260824_120000_filtered_students.xlsx")

    assert is_generated_export_filename(path)
    with pytest.raises(GeneratedExportWorkbookError, match="filtered export"):
        reject_generated_export_workbook(path)


def test_renamed_generated_export_is_rejected_by_sheet_signature(tmp_path: Path) -> None:
    path = tmp_path / "renamed.xlsx"
    workbook = Workbook()
    workbook.active.title = "Filtered Results"
    workbook.create_sheet("Filter Metadata")
    workbook.save(path)

    with pytest.raises(GeneratedExportWorkbookError, match="Filter Metadata"):
        reject_generated_export_workbook(path)


def test_file_size_stability_wait_accepts_stable_size(tmp_path: Path) -> None:
    path = create_minimal_workbook(tmp_path / "WSP.xlsx")
    sizes = deque([10, 20, 20])
    current_time = {"value": 0.0}

    def size_reader(_path: Path) -> int:
        return sizes.popleft()

    def sleep(_seconds: float) -> None:
        current_time["value"] += 0.1

    result = wait_for_file_size_to_stabilize(
        path,
        poll_interval_seconds=0.1,
        timeout_seconds=1,
        size_reader=size_reader,
        sleep_func=sleep,
        clock_func=lambda: current_time["value"],
    )

    assert result == 20


def test_unstable_file_times_out_predictably(tmp_path: Path) -> None:
    path = create_minimal_workbook(tmp_path / "WSP.xlsx")
    current_time = {"value": 0.0, "size": 0}

    def size_reader(_path: Path) -> int:
        current_time["size"] += 1
        return current_time["size"]

    def sleep(seconds: float) -> None:
        current_time["value"] += seconds

    with pytest.raises(FileNotStableError, match="did not stabilize"):
        wait_for_file_size_to_stabilize(
            path,
            poll_interval_seconds=0.1,
            timeout_seconds=0.25,
            size_reader=size_reader,
            sleep_func=sleep,
            clock_func=lambda: current_time["value"],
        )


def test_wait_for_file_size_to_stabilize_validates_settings(tmp_path: Path) -> None:
    path = create_minimal_workbook(tmp_path / "WSP.xlsx")

    with pytest.raises(ValueError, match="at least 2"):
        wait_for_file_size_to_stabilize(path, required_stable_reads=1)

    with pytest.raises(ValueError, match="cannot be negative"):
        wait_for_file_size_to_stabilize(path, poll_interval_seconds=-1)

    with pytest.raises(ValueError, match="cannot be negative"):
        wait_for_file_size_to_stabilize(path, timeout_seconds=-1)


def test_same_file_produces_same_hash(tmp_path: Path) -> None:
    path = create_minimal_workbook(tmp_path / "WSP.xlsx")

    first_hash = calculate_file_hash(path)
    second_hash = calculate_file_hash(path)

    assert first_hash == second_hash
    assert len(first_hash) == 64


def test_changed_file_produces_different_hash(tmp_path: Path) -> None:
    path = create_minimal_workbook(tmp_path / "WSP.xlsx")
    original_hash = calculate_file_hash(path)

    workbook = Workbook()
    workbook.active.append(["STUD_ID"])
    workbook.active.append(["1001"])
    workbook.save(path)

    assert calculate_file_hash(path) != original_hash


def test_calculate_file_hash_validates_chunk_size(tmp_path: Path) -> None:
    path = create_minimal_workbook(tmp_path / "WSP.xlsx")

    with pytest.raises(ValueError, match="chunk_size must be positive"):
        calculate_file_hash(path, chunk_size=0)


def test_duplicate_import_is_detected_and_logged(tmp_path: Path) -> None:
    engine = create_sqlite_engine(tmp_path / "wsp.db")
    initialize_database(engine)
    session_factory = create_session_factory(engine)
    workbook_path = create_minimal_workbook(tmp_path / "WSP.xlsx")
    file_hash = calculate_file_hash(workbook_path)

    with session_factory() as session:
        batch = ImportBatch(filename="WSP.xlsx", file_path=str(workbook_path), file_hash=file_hash, status="completed")
        session.add(batch)
        session.commit()

        with pytest.raises(DuplicateExcelFileError, match="already imported"):
            ensure_file_not_previously_imported(session, workbook_path, file_hash=file_hash)
        session.commit()

    with session_factory() as session:
        event = session.query(FileImportLog).filter_by(event_type="duplicate_file_skipped").one()

    assert event.details_json["file_hash"] == file_hash
    assert event.batch_id == batch.batch_id


def test_new_file_hash_is_returned_when_not_previously_imported(tmp_path: Path) -> None:
    engine = create_sqlite_engine(tmp_path / "wsp.db")
    initialize_database(engine)
    session_factory = create_session_factory(engine)
    workbook_path = create_minimal_workbook(tmp_path / "WSP.xlsx")

    with session_factory() as session:
        file_hash = ensure_file_not_previously_imported(session, workbook_path)

    assert file_hash == calculate_file_hash(workbook_path)


def test_find_import_batch_by_hash_returns_existing_batch(tmp_path: Path) -> None:
    engine = create_sqlite_engine(tmp_path / "wsp.db")
    initialize_database(engine)
    session_factory = create_session_factory(engine)

    with session_factory() as session:
        batch = ImportBatch(filename="WSP.xlsx", file_path="C:/WSP.xlsx", file_hash="hash-1")
        session.add(batch)
        session.commit()

        found = find_import_batch_by_hash(session, "hash-1")

    assert found is not None
    assert found.batch_id == batch.batch_id


def test_duplicate_skip_does_not_modify_students(tmp_path: Path) -> None:
    engine = create_sqlite_engine(tmp_path / "wsp.db")
    initialize_database(engine)
    session_factory = create_session_factory(engine)
    workbook_path = create_minimal_workbook(tmp_path / "WSP.xlsx")
    file_hash = calculate_file_hash(workbook_path)

    with session_factory() as session:
        session.add(ImportBatch(filename="WSP.xlsx", file_path=str(workbook_path), file_hash=file_hash))
        session.add(StudentCurrent(STUD_ID="1001", STUD_NAME="Existing"))
        session.commit()

        with pytest.raises(DuplicateExcelFileError):
            ensure_file_not_previously_imported(session, workbook_path, file_hash=file_hash)
        session.commit()

    with session_factory() as session:
        students = session.query(StudentCurrent).all()

    assert len(students) == 1
    assert students[0].STUD_NAME == "Existing"


def test_read_excel_workbook_loads_fixture(sample_workbook_path: Path) -> None:
    workbook = read_excel_workbook(sample_workbook_path)

    assert workbook.active_sheet == "Sheet1"
    assert workbook.row_count == 1
    assert workbook.column_count > 1
    assert workbook.rows[0]["STUD_ID"] == "1001"


def test_read_excel_workbook_detects_headers(sample_workbook_path: Path) -> None:
    workbook = read_excel_workbook(sample_workbook_path)

    assert workbook.headers[0] == "STUD_ID"
    assert "CUM_GPA" in workbook.headers


def test_read_excel_workbook_preserves_empty_cells_as_none(tmp_path: Path) -> None:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.append(["STUD_ID", "CUM_GPA", "NOTES"])
    worksheet.append(["1001", None, ""])
    path = tmp_path / "empty_cells.xlsx"
    workbook.save(path)

    data = read_excel_workbook(path)

    assert data.rows == ({"STUD_ID": "1001", "CUM_GPA": None, "NOTES": None},)


def test_read_excel_workbook_uses_first_sheet_by_default(tmp_path: Path) -> None:
    workbook = Workbook()
    first = workbook.active
    first.title = "First"
    first.append(["STUD_ID"])
    first.append(["first-student"])
    second = workbook.create_sheet("Second")
    second.append(["STUD_ID"])
    second.append(["second-student"])
    path = tmp_path / "multiple_sheets.xlsx"
    workbook.save(path)

    data = read_excel_workbook(path)

    assert data.sheet_names == ("First", "Second")
    assert data.active_sheet == "First"
    assert data.rows[0]["STUD_ID"] == "first-student"


def test_read_excel_workbook_allows_future_sheet_selection(tmp_path: Path) -> None:
    workbook = Workbook()
    first = workbook.active
    first.title = "First"
    first.append(["STUD_ID"])
    first.append(["first-student"])
    second = workbook.create_sheet("Second")
    second.append(["STUD_ID"])
    second.append(["second-student"])
    path = tmp_path / "multiple_sheets.xlsx"
    workbook.save(path)

    data = read_excel_workbook(path, sheet_name="Second")

    assert data.active_sheet == "Second"
    assert data.rows[0]["STUD_ID"] == "second-student"


def test_read_excel_workbook_rejects_missing_sheet(sample_workbook_path: Path) -> None:
    with pytest.raises(ValueError, match="was not found"):
        read_excel_workbook(sample_workbook_path, sheet_name="Missing")


def test_normalize_student_id_trims_and_normalizes_excel_numeric_values() -> None:
    assert normalize_student_id(" 1001 ") == "1001"
    assert normalize_student_id(1001) == "1001"
    assert normalize_student_id(1001.0) == "1001"
    assert normalize_student_id("1001.0") == "1001"


def test_normalize_student_id_rejects_empty_or_boolean_values() -> None:
    assert normalize_student_id(None) is None
    assert normalize_student_id("") is None
    assert normalize_student_id(True) is None


def test_extract_required_student_id_returns_valid_id() -> None:
    assert extract_required_student_id({"STUD_ID": " 1001 "}, row_number=2) == "1001"


def test_extract_required_student_id_rejects_and_logs_missing_id(tmp_path: Path) -> None:
    engine = create_sqlite_engine(tmp_path / "wsp.db")
    initialize_database(engine)
    session_factory = create_session_factory(engine)

    with session_factory() as session:
        batch = ImportBatch(filename="WSP.xlsx", file_path="C:/WSP.xlsx", file_hash="hash-1")
        session.add(batch)
        session.commit()

        with pytest.raises(MissingStudentIdError, match="usable STUD_ID"):
            extract_required_student_id(
                {"STUD_ID": "", "STUD_NAME": "No ID Student"},
                session=session,
                batch_id=batch.batch_id,
                row_number=7,
            )
        session.commit()

    with session_factory() as session:
        event = session.query(FileImportLog).filter_by(event_type="row_rejected").one()

    assert event.row_number == 7
    assert event.message == "Missing or invalid STUD_ID"
    assert event.details_json["reason"] == "missing_stud_id"
    assert event.details_json["row"]["STUD_NAME"] == "No ID Student"


def test_same_normalized_row_produces_same_hash() -> None:
    row = {"STUD_ID": "1001", "CUM_GPA": 3.4, "STUD_NAME": "Student"}

    assert generate_row_hash(row) == generate_row_hash(dict(row))


def test_changed_gpa_changes_row_hash() -> None:
    first = {"STUD_ID": "1001", "CUM_GPA": 3.4}
    second = {"STUD_ID": "1001", "CUM_GPA": 3.5}

    assert generate_row_hash(first) != generate_row_hash(second)


def test_changed_extra_column_changes_row_hash() -> None:
    first = {"STUD_ID": "1001", "extra_columns_json": {"NEW_COLUMN": "A"}}
    second = {"STUD_ID": "1001", "extra_columns_json": {"NEW_COLUMN": "B"}}

    assert generate_row_hash(first) != generate_row_hash(second)


def test_column_order_does_not_change_row_hash() -> None:
    first = {"STUD_ID": "1001", "CUM_GPA": 3.4, "STUD_NAME": "Student"}
    second = {"STUD_NAME": "Student", "STUD_ID": "1001", "CUM_GPA": 3.4}

    assert generate_row_hash(first) == generate_row_hash(second)


def test_volatile_database_fields_are_excluded_from_row_hash() -> None:
    first = {"STUD_ID": "1001", "CUM_GPA": 3.4, "updated_at": "2026-01-01", "row_hash": "old"}
    second = {"STUD_ID": "1001", "CUM_GPA": 3.4, "updated_at": "2026-06-04", "row_hash": "new"}

    assert generate_row_hash(first) == generate_row_hash(second)


def test_upsert_student_row_inserts_new_student(tmp_path: Path) -> None:
    engine = create_sqlite_engine(tmp_path / "wsp.db")
    initialize_database(engine)
    session_factory = create_session_factory(engine)

    with session_factory() as session:
        batch = ImportBatch(filename="WSP.xlsx", file_path="C:/WSP.xlsx", file_hash="hash-1")
        session.add(batch)
        session.commit()

        result = upsert_student_row(
            session,
            {"STUD_ID": "1001", "STUD_NAME": "New Student", "CUM_GPA": 3.4, "NEW_COLUMN": "extra"},
            batch_id=batch.batch_id,
        )
        session.commit()

    with session_factory() as session:
        student = session.query(StudentCurrent).filter_by(STUD_ID="1001").one()

    assert result.change_type == "new_student"
    assert student.STUD_NAME == "New Student"
    assert student.CUM_GPA == 3.4
    assert student.extra_columns_json == {"NEW_COLUMN": "extra"}
    assert student.first_seen_batch_id == batch.batch_id
    assert student.last_seen_batch_id == batch.batch_id
    assert student.added_to_db_at is not None
    assert student.modified_in_db_at is None


def test_upsert_student_row_marks_unchanged_without_history(tmp_path: Path) -> None:
    engine = create_sqlite_engine(tmp_path / "wsp.db")
    initialize_database(engine)
    session_factory = create_session_factory(engine)

    with session_factory() as session:
        first_batch = ImportBatch(filename="WSP.xlsx", file_path="C:/WSP.xlsx", file_hash="hash-1")
        second_batch = ImportBatch(filename="WSP2.xlsx", file_path="C:/WSP2.xlsx", file_hash="hash-2")
        session.add_all([first_batch, second_batch])
        session.commit()

        row = {"STUD_ID": "1001", "STUD_NAME": "Same Student", "CUM_GPA": 3.4}
        upsert_student_row(session, row, batch_id=first_batch.batch_id)
        result = upsert_student_row(session, row, batch_id=second_batch.batch_id)
        session.commit()

    with session_factory() as session:
        student = session.query(StudentCurrent).filter_by(STUD_ID="1001").one()
        history_count = session.query(StudentHistory).count()

    assert result.change_type == "unchanged_student"
    assert student.last_seen_batch_id == second_batch.batch_id
    assert student.modified_in_db_at is None
    assert history_count == 0


def test_upsert_student_row_updates_changed_student_and_creates_history(tmp_path: Path) -> None:
    engine = create_sqlite_engine(tmp_path / "wsp.db")
    initialize_database(engine)
    session_factory = create_session_factory(engine)

    with session_factory() as session:
        first_batch = ImportBatch(filename="WSP.xlsx", file_path="C:/WSP.xlsx", file_hash="hash-1")
        second_batch = ImportBatch(filename="WSP2.xlsx", file_path="C:/WSP2.xlsx", file_hash="hash-2")
        session.add_all([first_batch, second_batch])
        session.commit()

        upsert_student_row(session, {"STUD_ID": "1001", "STUD_NAME": "Student", "CUM_GPA": 3.4}, batch_id=first_batch.batch_id)
        result = upsert_student_row(session, {"STUD_ID": "1001", "STUD_NAME": "Student", "CUM_GPA": 3.8}, batch_id=second_batch.batch_id)
        session.commit()

    with session_factory() as session:
        student = session.query(StudentCurrent).filter_by(STUD_ID="1001").one()
        history = session.query(StudentHistory).filter_by(STUD_ID="1001").one()

    assert result.change_type == "updated_student"
    assert student.CUM_GPA == 3.8
    assert student.last_seen_batch_id == second_batch.batch_id
    assert student.modified_in_db_at is not None
    assert history.all_excel_columns["CUM_GPA"] == 3.4
    assert history.change_type == "updated_student"


def test_upsert_student_row_clears_missing_flag_for_restored_student(tmp_path: Path) -> None:
    engine = create_sqlite_engine(tmp_path / "wsp.db")
    initialize_database(engine)
    session_factory = create_session_factory(engine)

    with session_factory() as session:
        first_batch = ImportBatch(filename="WSP.xlsx", file_path="C:/WSP.xlsx", file_hash="hash-1")
        second_batch = ImportBatch(filename="WSP2.xlsx", file_path="C:/WSP2.xlsx", file_hash="hash-2")
        session.add_all([first_batch, second_batch])
        session.commit()

        upsert_student_row(session, {"STUD_ID": "1001", "STUD_NAME": "Student"}, batch_id=first_batch.batch_id)
        student = session.query(StudentCurrent).filter_by(STUD_ID="1001").one()
        student.missing_from_latest_import = True
        session.commit()

        result = upsert_student_row(session, {"STUD_ID": "1001", "STUD_NAME": "Student"}, batch_id=second_batch.batch_id)
        session.commit()

    with session_factory() as session:
        student = session.query(StudentCurrent).filter_by(STUD_ID="1001").one()

    assert result.change_type == "restored_student"
    assert student.missing_from_latest_import is False


def test_mark_missing_students_marks_absent_student_without_deleting(tmp_path: Path) -> None:
    engine = create_sqlite_engine(tmp_path / "wsp.db")
    initialize_database(engine)
    session_factory = create_session_factory(engine)

    with session_factory() as session:
        batch = ImportBatch(filename="WSP.xlsx", file_path="C:/WSP.xlsx", file_hash="hash-1")
        session.add(batch)
        session.commit()
        upsert_student_row(session, {"STUD_ID": "1001", "STUD_NAME": "Present"}, batch_id=batch.batch_id)
        upsert_student_row(session, {"STUD_ID": "1002", "STUD_NAME": "Missing"}, batch_id=batch.batch_id)
        result = mark_missing_students(session, seen_student_ids={"1001"}, batch_id=batch.batch_id)
        session.commit()

    with session_factory() as session:
        students = {student.STUD_ID: student for student in session.query(StudentCurrent).all()}
        history = session.query(StudentHistory).filter_by(STUD_ID="1002").one()

    assert result.newly_missing_student_ids == ("1002",)
    assert students["1001"].missing_from_latest_import is False
    assert students["1002"].missing_from_latest_import is True
    assert history.change_type == "missing_from_latest_import"


def test_mark_missing_students_does_not_duplicate_history_for_already_missing_student(tmp_path: Path) -> None:
    engine = create_sqlite_engine(tmp_path / "wsp.db")
    initialize_database(engine)
    session_factory = create_session_factory(engine)

    with session_factory() as session:
        batch = ImportBatch(filename="WSP.xlsx", file_path="C:/WSP.xlsx", file_hash="hash-1")
        session.add(batch)
        session.commit()
        upsert_student_row(session, {"STUD_ID": "1001", "STUD_NAME": "Missing"}, batch_id=batch.batch_id)
        first = mark_missing_students(session, seen_student_ids=set(), batch_id=batch.batch_id)
        second = mark_missing_students(session, seen_student_ids=set(), batch_id=batch.batch_id)
        session.commit()

    with session_factory() as session:
        history_count = session.query(StudentHistory).filter_by(STUD_ID="1001").count()

    assert first.newly_missing_student_ids == ("1001",)
    assert second.newly_missing_student_ids == ()
    assert second.already_missing_student_ids == ("1001",)
    assert history_count == 1


def test_mark_missing_students_leaves_seen_students_active(tmp_path: Path) -> None:
    engine = create_sqlite_engine(tmp_path / "wsp.db")
    initialize_database(engine)
    session_factory = create_session_factory(engine)

    with session_factory() as session:
        batch = ImportBatch(filename="WSP.xlsx", file_path="C:/WSP.xlsx", file_hash="hash-1")
        session.add(batch)
        session.commit()
        upsert_student_row(session, {"STUD_ID": "1001", "STUD_NAME": "Present"}, batch_id=batch.batch_id)
        result = mark_missing_students(session, seen_student_ids={"1001"}, batch_id=batch.batch_id)
        session.commit()

    with session_factory() as session:
        student = session.query(StudentCurrent).filter_by(STUD_ID="1001").one()

    assert result.newly_missing_student_ids == ()
    assert student.missing_from_latest_import is False


def test_execute_import_transaction_commits_successful_import(tmp_path: Path) -> None:
    engine = create_sqlite_engine(tmp_path / "wsp.db")
    initialize_database(engine)
    session_factory = create_session_factory(engine)

    with session_factory() as session:
        batch = ImportBatch(filename="WSP.xlsx", file_path="C:/WSP.xlsx", file_hash="hash-1")
        session.add(batch)
        session.commit()
        batch_id = batch.batch_id

    def operation(session, batch):
        upsert_student_row(session, {"STUD_ID": "1001", "STUD_NAME": "Student"}, batch_id=batch.batch_id)
        return "ok"

    result = execute_import_transaction(session_factory, batch_id=batch_id, operation=operation)

    with session_factory() as session:
        batch = session.get(ImportBatch, batch_id)
        student_count = session.query(StudentCurrent).count()

    assert result == "ok"
    assert batch.status == "completed"
    assert student_count == 1


def test_execute_import_transaction_rolls_back_student_changes_and_records_failed_batch(tmp_path: Path) -> None:
    engine = create_sqlite_engine(tmp_path / "wsp.db")
    initialize_database(engine)
    session_factory = create_session_factory(engine)
    archived_file = tmp_path / "archive" / "WSP.xlsx"
    archived_file.parent.mkdir()
    archived_file.write_text("archived copy", encoding="utf-8")

    with session_factory() as session:
        batch = ImportBatch(
            filename="WSP.xlsx",
            file_path="C:/WSP.xlsx",
            archived_file_path=str(archived_file),
            file_hash="hash-1",
        )
        session.add(batch)
        session.commit()
        batch_id = batch.batch_id

    def operation(session, batch):
        upsert_student_row(session, {"STUD_ID": "1001", "STUD_NAME": "Partial"}, batch_id=batch.batch_id)
        raise RuntimeError("simulated import failure")

    with pytest.raises(RuntimeError, match="simulated import failure"):
        execute_import_transaction(session_factory, batch_id=batch_id, operation=operation)

    with session_factory() as session:
        batch = session.get(ImportBatch, batch_id)
        student_count = session.query(StudentCurrent).count()

    assert student_count == 0
    assert batch.status == "failed"
    assert batch.error_message == "simulated import failure"
    assert archived_file.exists()


def test_execute_import_transaction_records_missing_batch_error_without_partial_rows(tmp_path: Path) -> None:
    engine = create_sqlite_engine(tmp_path / "wsp.db")
    initialize_database(engine)
    session_factory = create_session_factory(engine)

    def operation(session, batch):
        upsert_student_row(session, {"STUD_ID": "1001"}, batch_id=batch.batch_id)

    with pytest.raises(ValueError, match="does not exist"):
        execute_import_transaction(session_factory, batch_id=999, operation=operation)

    with session_factory() as session:
        assert session.query(StudentCurrent).count() == 0
