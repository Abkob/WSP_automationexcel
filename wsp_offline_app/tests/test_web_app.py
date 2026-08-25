from __future__ import annotations

import os
from pathlib import Path
from shutil import copy2

from fastapi.testclient import TestClient
from openpyxl import Workbook

from app.web_app import build_filter_request_from_payload, create_web_app
from config import AppSettings
from database.db import create_session_factory, create_sqlite_engine, initialize_database
from database.models import BackupLog, ColumnRegistry, FileImportLog, ImportBatch, StudentCurrent, StudentHistory
from services.excel_schema import EXPECTED_WSP_COLUMNS


def make_test_client(tmp_path: Path) -> tuple[TestClient, AppSettings]:
    settings = AppSettings(data_dir=tmp_path / "data", runtime_mode="testing")
    engine = create_sqlite_engine(settings.database_path)
    initialize_database(engine)
    session_factory = create_session_factory(engine)
    with session_factory() as session:
        session.add(ImportBatch(filename="dummy.xlsx", file_path="dummy.xlsx", file_hash="hash-1", status="completed", new_rows=3))
        session.add(BackupLog(backup_path=str(settings.backup_dir / "dummy.db"), reason="test", status="created", integrity_check_passed=True))
        session.add(ColumnRegistry(column_name="STUD_ID", original_column_name="STUD_ID", detected_type="text", is_active=True))
        session.add(ColumnRegistry(column_name="CUM_GPA", original_column_name="CUM_GPA", detected_type="number", is_active=True))
        session.add(FileImportLog(event_type="normalization_warning", message="CUM_GPA: Invalid numeric value", row_number=4))
        session.add_all(
            [
                StudentCurrent(
                    STUD_ID="260201",
                    STUD_NAME="Maya Rerun",
                    MAJR_DESC="Information Technology",
                    CLAS_DESC="Junior",
                    CUM_GPA=3.72,
                    PROBATION=False,
                    FINANCIAL_AID=True,
                    DORMS=False,
                    WSP_TECHNICAL_SKILLS="Excel, SQL, spreadsheet QA, dashboard notes",
                    WSP_PREFERRED_TYPE_OF_WORK="Data entry, spreadsheet reporting, and checking import errors.",
                ),
                StudentCurrent(
                    STUD_ID="260202",
                    STUD_NAME="Omar Rerun",
                    MAJR_DESC="Business Administration",
                    CLAS_DESC="Senior",
                    CUM_GPA=3.41,
                    PROBATION=False,
                    FINANCIAL_AID=True,
                    DORMS=False,
                    WSP_TECHNICAL_SKILLS="PowerPoint and invoice tracking",
                    WSP_PREFERRED_TYPE_OF_WORK="Office assistant work with budgets and forms.",
                ),
                StudentCurrent(
                    STUD_ID="260203",
                    STUD_NAME="Nadine Lab",
                    MAJR_DESC="Biology",
                    CLAS_DESC="Sophomore",
                    CUM_GPA=3.1,
                    PROBATION=False,
                    FINANCIAL_AID=False,
                    DORMS=True,
                    WSP_TECHNICAL_SKILLS="Lab logs and sample labels",
                    WSP_PREFERRED_TYPE_OF_WORK="Laboratory assistant work.",
                ),
            ]
        )
        session.commit()
    return TestClient(create_web_app(settings)), settings


def test_production_app_prewarms_embeddings_and_registers_import_refresher(tmp_path: Path, monkeypatch) -> None:
    settings = AppSettings(data_dir=tmp_path / "production-data", runtime_mode="production")
    warmed = []
    monkeypatch.setattr("app.web_app._prewarm_embedding_model", lambda received: warmed.append(received))

    app = create_web_app(settings)

    assert warmed == [settings]
    handler_names = {getattr(handler, "__name__", "") for handler in app.router.on_startup}
    assert "start_import_folder_refresher" in handler_names


def test_filters_page_uses_fastapi_static_ui_not_nicegui(tmp_path: Path) -> None:
    client, _settings = make_test_client(tmp_path)

    response = client.get("/filters")

    assert response.status_code == 200
    assert "Filter Builder" in response.text
    assert "American University of Beirut" in response.text
    assert "AI semantic matching" in response.text
    assert 'id="ai-search-progress"' in response.text
    assert "Searching the local AI index" in response.text
    loader_css = client.get("/static/wsp.css")
    assert loader_css.status_code == 200
    assert "@keyframes aubDotPulse" in loader_css.text
    assert "linear-gradient(90deg, var(--aub), var(--aub-2), var(--aub))" in loader_css.text
    assert "/static/wsp.css" in response.text
    assert "_nicegui" not in response.text


def test_index_status_uses_active_students_as_coverage_denominator(tmp_path: Path) -> None:
    client, settings = make_test_client(tmp_path)
    session_factory = create_session_factory(create_sqlite_engine(settings.database_path))
    with session_factory() as session:
        session.add(StudentCurrent(STUD_ID="inactive", missing_from_latest_import=True))
        session.commit()

    response = client.get("/api/admin/index-status")

    assert response.status_code == 200
    payload = response.json()
    assert payload["db_count"] == 3
    assert payload["total_db_count"] == 4


def test_admin_workspace_pages_render_from_fastapi_shell(tmp_path: Path) -> None:
    client, _settings = make_test_client(tmp_path)

    pages = {
        "/excel-sheets": "Excel Sheets",
        "/student-profile": "Student Profile",
        "/import": "Import Center",
        "/system-status": "System Health",
    }

    for path, heading in pages.items():
        response = client.get(path)
        assert response.status_code == 200
        assert heading in response.text
        assert "/static/wsp.css" in response.text
        assert f'body data-active-path="{path}"' in response.text
        assert "_nicegui" not in response.text

    import_page = client.get("/import")
    assert 'href="/excel-sheets"' in import_page.text
    assert 'href="/import"' in import_page.text
    assert 'href="/system-status"' in import_page.text
    assert "Check Now" in import_page.text
    assert "Save Folder" in import_page.text
    assert "Auto-consume active" in import_page.text
    assert "disabled><span" not in import_page.text

    sheets_page = client.get("/excel-sheets")
    assert "Database worksheet views" in sheets_page.text
    assert "sheet-source-map" in sheets_page.text

    filters_page = client.get("/filters")
    assert "Filtering source" in filters_page.text
    assert "students_current" in filters_page.text


def test_student_profile_page_and_api_present_complete_record(tmp_path: Path) -> None:
    client, settings = make_test_client(tmp_path)
    session_factory = create_session_factory(create_sqlite_engine(settings.database_path))
    with session_factory() as session:
        student = session.query(StudentCurrent).filter_by(STUD_ID="260201").one()
        student.STUD_EMAIL = "maya.rerun@aub.edu.lb"
        student.MOBILE_NBR = "+961 70 000 201"
        student.TOTAL_CREDIT_HOURS = 74
        student.ENROLLED_IND = True
        student.REGISTERED_IND = True
        student.ASTD_DESC = "Good Standing"
        student.WSP_ORGANIZATIONAL_SKILLS = "Scheduling; documentation"
        student.WSP_SPOKEN_LANGUAGES = "Arabic, English"
        student.extra_columns_json = {"SUPERVISOR_NAME": "Dr. Example", "OFFICE_LOCATION": "West Hall"}
        session.add(
            StudentHistory(
                batch_id=1,
                STUD_ID="260201",
                all_excel_columns={"CUM_GPA": 3.6},
                change_type="updated_student",
            )
        )
        session.commit()

    page = client.get("/student-profile/260201")
    assert page.status_code == 200
    assert 'data-active-path="/student-profile"' in page.text
    assert 'data-student-id="260201"' in page.text
    assert "Find a student" in page.text
    assert "Student Profiles" in page.text

    response = client.get("/api/students/260201/profile")
    assert response.status_code == 200
    payload = response.json()
    assert payload["identity"]["name"] == "Maya Rerun"
    assert payload["identity"]["email"] == "maya.rerun@aub.edu.lb"
    assert payload["highlights"][0]["value"] == "3.72"
    assert payload["skills"][0]["values"][:2] == ["Excel", "SQL"]
    assert payload["support"][0] == {"label": "Financial aid", "field": "FINANCIAL_AID", "value": True}
    assert payload["additional_fields"][1]["label"] == "Supervisor Name"
    assert payload["timeline"][0]["title"] in {"Student record updated", "Student record added"}


def test_student_lookup_and_missing_profile(tmp_path: Path) -> None:
    client, _settings = make_test_client(tmp_path)

    lookup = client.get("/api/students/lookup", params={"q": "Maya", "limit": 5})
    assert lookup.status_code == 200
    assert lookup.json()["students"][0]["student_id"] == "260201"

    response = client.get("/api/students/does-not-exist/profile")
    assert response.status_code == 404
    assert response.json()["detail"] == "Student profile not found"


def test_filter_and_excel_assets_include_profile_entry_points(tmp_path: Path) -> None:
    client, _settings = make_test_client(tmp_path)

    script = client.get("/static/wsp.js").text
    styles = client.get("/static/wsp.css").text
    assert "student-profile-link" in script
    assert "bindStudentProfileActions" in script
    assert 'addEventListener("contextmenu"' in script
    assert ".student-context-menu" in styles
    assert ".profile-hero" in styles


def test_excel_global_search_renderer_defines_edit_mode(tmp_path: Path) -> None:
    client, _settings = make_test_client(tmp_path)

    script = client.get("/static/wsp.js").text
    global_renderer = script.split("function renderActiveSheetWithQuery", 1)[1].split(
        "function renderActiveSheet()", 1
    )[0]

    assert 'const editMode = sheet.key === "Student_Directory" && sheetState.editMode;' in global_renderer


def test_dashboard_api_returns_metrics_and_chart_data(tmp_path: Path) -> None:
    client, _settings = make_test_client(tmp_path)

    response = client.get("/api/dashboard")

    assert response.status_code == 200
    payload = response.json()
    assert payload["metrics"]["total_students"] == 3
    assert payload["latest_import"]["filename"] == "dummy.xlsx"
    assert payload["charts"]["students_by_major"]
    assert payload["filter_options"]["faculties"]
    assert payload["faculty_summary"]
    assert payload["total_matches"] == 3
    assert payload["preferred_work_grouping"]["method"] in {"offline_embeddings", "review_fallback"}
    assert payload["preferred_work_grouping"]["original_text_preserved"] is True
    assert payload["preferred_work_grouping"]["preference_count"] == 3
    assert payload["charts"]["work_preferences"]
    assert payload["students"][0]["work_preference_group"]


def test_dashboard_api_supports_faculty_and_class_drilldown(tmp_path: Path) -> None:
    client, _settings = make_test_client(tmp_path)

    response = client.get("/api/dashboard", params={"faculty": "OSB", "class_year": "Senior"})

    assert response.status_code == 200
    payload = response.json()
    assert payload["selection_label"] == "OSB · Senior"
    assert payload["total_matches"] == 1
    assert payload["students"][0]["student_id"] == "260202"
    assert payload["students"][0]["faculty"] == "OSB"


def test_dashboard_static_assets_include_interactive_chart_system(tmp_path: Path) -> None:
    client, _settings = make_test_client(tmp_path)

    script = client.get("/static/wsp.js")
    styles = client.get("/static/wsp.css")

    assert script.status_code == 200
    assert styles.status_code == 200
    assert "renderChartCard" in script.text
    assert "renderChartGrid" in script.text
    assert "renderChartStatPill" in script.text
    assert "renderSinglePointSpotlight" in script.text
    assert "initDashboardMultiSelect" in script.text
    assert "renderDashboardFacultyComparison" in script.text
    assert "percent-bars" in script.text
    assert "Candidate context by faculty" in script.text
    assert "Ready for matching" not in script.text
    assert "Place the right student in the right role." not in client.get("/").text
    assert "Grouped locally with offline embeddings" in script.text
    assert "work_preference_group" in script.text
    assert "handleDashboardChartSelection" in script.text
    assert "loadExcelSheets" in script.text
    assert "loadImportCenter" in script.text
    assert "loadSystemStatus" in script.text
    assert "refreshUploadFolder" in script.text
    assert "saveImportFolder" in script.text
    assert ".chart-card" in styles.text
    assert ".chart-canvas-wrap" in styles.text
    assert ".chart-head" in styles.text
    assert ".dashboard-page-head" in styles.text
    assert ".dashboard-faculty-card" in styles.text
    assert ".dashboard-grouping-note" in styles.text
    assert ".chart-single-spotlight" in styles.text
    assert 'id="dashboard-ms-major"' in client.get("/").text
    assert ".dashboard-work-group" in styles.text
    assert ".sheet-workspace" in styles.text
    assert ".system-info-panel" in styles.text
    assert ".source-card" in styles.text
    assert ".policy-grid" in styles.text
    assert ".folder-summary" in styles.text


def test_new_workspace_apis_return_backend_data(tmp_path: Path) -> None:
    client, _settings = make_test_client(tmp_path)

    sheets = client.get("/api/excel-sheets")
    import_center = client.get("/api/import-center")
    system_status = client.get("/api/system-status")

    assert sheets.status_code == 200
    sheet_payload = sheets.json()
    assert sheet_payload["source_map"][0]["label"] == "Import Folder"
    assert sheet_payload["source_map"][2]["value"] == "SQLite students_current"
    assert {sheet["key"] for sheet in sheet_payload["sheets"]} == {
        "Student_Directory",
        "Skipped_Rows",
        "Column_Schema",
        "Major_Analytics",
    }
    assert sheet_payload["sheets"][0]["rows"][0][0] == "260201"

    assert import_center.status_code == 200
    import_payload = import_center.json()
    assert import_payload["paths"]["import_folder"]
    assert import_payload["paths"]["upload_folder"] == import_payload["paths"]["watched_folder"]
    assert import_payload["import_folder"]["archive_folder"] == import_payload["paths"]["archive_folder"]
    assert import_payload["backup_policy"]["pre_import"] == "Created before the database is changed."
    assert import_payload["auto_refresh"]["enabled"] is True
    assert import_payload["recent_logs"][0]["event_type"] == "normalization_warning"
    assert import_payload["columns"][0]["column_name"] in {"CUM_GPA", "STUD_ID"}

    assert system_status.status_code == 200
    status_payload = system_status.json()
    assert "health" in status_payload
    assert "system" in status_payload


def test_import_run_api_executes_safe_excel_pipeline(tmp_path: Path, sample_workbook_path: Path) -> None:
    client, settings = make_test_client(tmp_path)

    response = client.post("/api/import/run", json={"path": str(sample_workbook_path)})

    assert response.status_code == 200
    payload = response.json()
    assert payload["status"] == "completed"
    assert payload["rows"] == 1
    assert payload["new_rows"] == 1
    assert payload["columns"] == 39
    assert settings.backup_dir.exists()
    assert settings.original_excel_archive_dir.exists()

    import_center = client.get("/api/import-center").json()
    assert import_center["latest_import"]["filename"] == sample_workbook_path.name
    assert len(import_center["recent_backups"]) >= 2


def test_import_run_rejects_headers_only_workbook_without_marking_students_missing(tmp_path: Path) -> None:
    client, settings = make_test_client(tmp_path)
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Sheet1"
    worksheet.append(EXPECTED_WSP_COLUMNS)
    path = tmp_path / "headers_only_wsp.xlsx"
    workbook.save(path)

    response = client.post("/api/import/run", json={"path": str(path)})

    assert response.status_code == 400
    assert "has headers but no student rows" in response.json()["detail"]
    dashboard = client.get("/api/dashboard").json()
    assert dashboard["metrics"]["total_students"] == 3

    engine = create_sqlite_engine(settings.database_path)
    session_factory = create_session_factory(engine)
    with session_factory() as session:
        assert session.query(StudentCurrent).filter(StudentCurrent.missing_from_latest_import.is_(False)).count() == 3
        failed_batch = session.query(ImportBatch).order_by(ImportBatch.batch_id.desc()).first()
        assert failed_batch.filename == "headers_only_wsp.xlsx"
        assert failed_batch.status == "failed"


def test_import_run_rejects_workbook_with_no_valid_student_ids_without_marking_students_missing(tmp_path: Path) -> None:
    client, settings = make_test_client(tmp_path)
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Sheet1"
    worksheet.append(EXPECTED_WSP_COLUMNS)
    invalid_row = [None for _column in EXPECTED_WSP_COLUMNS]
    invalid_row[EXPECTED_WSP_COLUMNS.index("STUD_NAME")] = "No ID Student"
    worksheet.append(invalid_row)
    path = tmp_path / "no_valid_ids_wsp.xlsx"
    workbook.save(path)

    response = client.post("/api/import/run", json={"path": str(path)})

    assert response.status_code == 400
    assert "none had a usable STUD_ID" in response.json()["detail"]
    dashboard = client.get("/api/dashboard").json()
    assert dashboard["metrics"]["total_students"] == 3

    engine = create_sqlite_engine(settings.database_path)
    session_factory = create_session_factory(engine)
    with session_factory() as session:
        assert session.query(StudentCurrent).filter(StudentCurrent.missing_from_latest_import.is_(False)).count() == 3
        failed_batch = session.query(ImportBatch).order_by(ImportBatch.batch_id.desc()).first()
        assert failed_batch.filename == "no_valid_ids_wsp.xlsx"
        assert failed_batch.status == "failed"


def test_import_refresh_folder_imports_new_uploads_and_skips_duplicates(tmp_path: Path, sample_workbook_path: Path) -> None:
    client, settings = make_test_client(tmp_path)
    settings.incoming_excel_dir.mkdir(parents=True, exist_ok=True)
    uploaded = settings.incoming_excel_dir / "uploaded_sample.xlsx"
    copy2(sample_workbook_path, uploaded)

    first = client.post("/api/import/refresh-folder")

    assert first.status_code == 200
    first_payload = first.json()
    assert first_payload["checked"] == 1
    assert first_payload["imported"] == 1
    assert first_payload["skipped"] == 0
    assert first_payload["results"][0]["filename"] == "uploaded_sample.xlsx"

    second = client.post("/api/import/refresh-folder")

    assert second.status_code == 200
    second_payload = second.json()
    assert second_payload["checked"] == 1
    assert second_payload["imported"] == 0
    assert second_payload["skipped"] == 1


def test_import_folder_can_be_assigned_from_import_center(tmp_path: Path) -> None:
    client, _settings = make_test_client(tmp_path)
    assigned = tmp_path / "Daily Import Folder"

    response = client.post("/api/import-folder", json={"path": str(assigned)})

    assert response.status_code == 200
    payload = response.json()
    assert payload["folder"] == str(assigned.resolve())
    assert (assigned / "archive").is_dir()

    import_center = client.get("/api/import-center").json()
    assert import_center["import_folder"]["path"] == str(assigned.resolve())
    assert import_center["paths"]["archive_folder"] == str((assigned / "archive").resolve())


def test_import_folder_keeps_newest_workbook_and_archives_old_root_files(tmp_path: Path, sample_workbook_path: Path) -> None:
    client, settings = make_test_client(tmp_path)
    settings.incoming_excel_dir.mkdir(parents=True, exist_ok=True)
    old_workbook = settings.incoming_excel_dir / "old_daily.xlsx"
    new_workbook = settings.incoming_excel_dir / "new_daily.xlsx"
    copy2(sample_workbook_path, old_workbook)

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Sheet1"
    worksheet.append(EXPECTED_WSP_COLUMNS)
    row = [None for _column in EXPECTED_WSP_COLUMNS]
    row[EXPECTED_WSP_COLUMNS.index("STUD_ID")] = "2002"
    row[EXPECTED_WSP_COLUMNS.index("STUD_NAME")] = "New Daily Student"
    row[EXPECTED_WSP_COLUMNS.index("CUM_GPA")] = 3.9
    worksheet.append(row)
    workbook.save(new_workbook)
    os.utime(old_workbook, (1_700_000_000, 1_700_000_000))
    os.utime(new_workbook, (1_700_000_100, 1_700_000_100))

    response = client.post("/api/import/refresh-folder")

    assert response.status_code == 200
    payload = response.json()
    assert payload["imported"] == 2
    assert payload["active_file"] == str(new_workbook.resolve())
    assert len(payload["archived_files"]) == 1
    assert not old_workbook.exists()
    assert new_workbook.exists()
    assert (settings.incoming_excel_dir / "archive").is_dir()
    assert Path(payload["archived_files"][0]["archived_to"]).exists()


def test_search_api_returns_fast_offline_semantic_results(tmp_path: Path) -> None:
    client, _settings = make_test_client(tmp_path)

    response = client.post(
        "/api/search",
        json={
            "semantic_query": "spreadsheet reporting with careful data entry",
            "name_query": "Rerun",
            "semantic_threshold": 0.5,
            "page_size": 10,
        },
    )

    assert response.status_code == 200
    payload = response.json()
    assert payload["total_count"] == 1
    assert payload["rows"][0]["STUD_ID"] == "260201"
    assert float(payload["rows"][0]["semantic_score"]) > 0.5
    assert payload["rows"][0]["semantic_explanation"]
    if "embedding search was unavailable" in payload["rows"][0]["semantic_explanation"]:
        assert payload["rows"][0]["semantic_explanation"].startswith("Text match fallback")


def test_export_api_creates_filtered_xlsx(tmp_path: Path) -> None:
    client, _settings = make_test_client(tmp_path)

    response = client.post(
        "/api/export",
        json={
            "semantic_query": "spreadsheet reporting",
            "name_query": "Rerun",
            "semantic_threshold": 0.5,
        },
    )

    assert response.status_code == 200
    payload = response.json()
    assert payload["row_count"] == 1
    assert Path(payload["path"]).exists()
    assert Path(payload["path"]).suffix == ".xlsx"


def test_web_payload_builder_handles_boolean_numeric_and_category_filters() -> None:
    request = build_filter_request_from_payload(
        {
            "gpa_min": "3.0",
            "gpa_max": "3.9",
            "probation": "no",
            "financial_aid": "yes",
            "major": "Information Technology",
            "semantic_query": "spreadsheet reporting",
            "semantic_threshold": "0.35",
            "page_size": "12",
            "sort_field": "CUM_GPA",
            "sort_direction": "desc",
        }
    )

    assert request.numeric_filters[0].operator == "between"
    assert request.boolean_filters[0].field_name == "PROBATION"
    assert request.boolean_filters[0].value is False
    assert request.boolean_filters[1].field_name == "FINANCIAL_AID"
    assert request.category_filters[0].values == ("Information Technology",)
    assert request.semantic_filter is not None
    assert request.semantic_filter.minimum_score == 0.35
    assert request.semantic_filter.top_k == 12
    assert request.sort is not None
    assert request.sort.field_name == "CUM_GPA"
