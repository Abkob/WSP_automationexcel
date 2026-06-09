from __future__ import annotations

from datetime import UTC, datetime
from pathlib import Path

import pytest
from openpyxl import load_workbook

from database.db import create_session_factory, create_sqlite_engine, initialize_database
from database.models import ExportLog, ImportBatch, StudentCurrent
from services.export_service import (
    build_export_filename,
    export_filtered_results_to_excel,
    safe_export_filename_part,
)
from services.filter_service import FilterRequest, FilterResult, TextFilter, execute_filter_request
from services.filter_service import log_filter_run


def test_safe_export_filename_part_removes_unsafe_characters() -> None:
    assert safe_export_filename_part(" High GPA / Python ") == "High_GPA_Python"
    assert safe_export_filename_part("...") == "filtered_results"


def test_build_export_filename_includes_timestamp_and_prefix() -> None:
    exported_at = datetime(2026, 6, 4, 14, 1, 5, tzinfo=UTC)

    assert build_export_filename("high gpa", exported_at) == "20260604_140105_high_gpa.xlsx"


def test_export_filtered_results_file_is_created(tmp_path: Path) -> None:
    result = FilterResult(
        rows=(),
        selected_rows=({"STUD_ID": "1001", "STUD_NAME": "Student"},),
        total_count=1,
        page=1,
        page_size=50,
        applied_filter_count=0,
        applied_filter_metadata={},
    )

    export = export_filtered_results_to_excel(
        result,
        tmp_path,
        exported_at=datetime(2026, 6, 4, 14, 1, 5, tzinfo=UTC),
    )

    assert export.path.exists()
    assert export.row_count == 1
    assert export.column_names == ("STUD_ID", "STUD_NAME")


def test_exported_workbook_opens_and_has_filtered_results_sheet(tmp_path: Path) -> None:
    result = FilterResult(
        rows=(),
        selected_rows=({"STUD_ID": "1001", "STUD_NAME": "Student"},),
        total_count=1,
        page=1,
        page_size=50,
        applied_filter_count=0,
        applied_filter_metadata={},
    )

    export = export_filtered_results_to_excel(result, tmp_path)
    workbook = load_workbook(export.path)

    assert "Filtered Results" in workbook.sheetnames
    worksheet = workbook["Filtered Results"]
    assert worksheet.freeze_panes == "A2"
    assert worksheet["A1"].font.bold is True


def test_exported_row_count_matches_filter_result(tmp_path: Path) -> None:
    result = FilterResult(
        rows=(),
        selected_rows=(
            {"STUD_ID": "1001", "STUD_NAME": "First"},
            {"STUD_ID": "1002", "STUD_NAME": "Second"},
        ),
        total_count=2,
        page=1,
        page_size=50,
        applied_filter_count=0,
        applied_filter_metadata={},
    )

    export = export_filtered_results_to_excel(result, tmp_path)
    workbook = load_workbook(export.path)
    worksheet = workbook["Filtered Results"]

    assert export.row_count == 2
    assert worksheet.max_row == 3


def test_export_respects_selected_columns(tmp_path: Path) -> None:
    engine = create_sqlite_engine(tmp_path / "wsp.db")
    initialize_database(engine)
    session_factory = create_session_factory(engine)

    with session_factory() as session:
        session.add(StudentCurrent(STUD_ID="1001", STUD_NAME="Alice", CUM_GPA=3.8, WSP_TECHNICAL_SKILLS="Python"))
        session.commit()

    with session_factory() as session:
        filter_result = execute_filter_request(
            session,
            FilterRequest(
                text_filters=(TextFilter("WSP_TECHNICAL_SKILLS", "contains", "Python"),),
                selected_columns=("STUD_ID", "STUD_NAME", "CUM_GPA"),
            ),
        )

    export = export_filtered_results_to_excel(filter_result, tmp_path)
    worksheet = load_workbook(export.path)["Filtered Results"]

    assert [cell.value for cell in worksheet[1]] == ["STUD_ID", "STUD_NAME", "CUM_GPA"]
    assert [cell.value for cell in worksheet[2]] == ["1001", "Alice", 3.8]


def test_export_includes_semantic_score_when_present(tmp_path: Path) -> None:
    result = FilterResult(
        rows=(),
        selected_rows=({"STUD_ID": "1001", "semantic_score": 0.91},),
        total_count=1,
        page=1,
        page_size=50,
        applied_filter_count=1,
        applied_filter_metadata={},
    )

    export = export_filtered_results_to_excel(result, tmp_path)
    worksheet = load_workbook(export.path)["Filtered Results"]

    assert [cell.value for cell in worksheet[1]] == ["STUD_ID", "semantic_score"]
    assert [cell.value for cell in worksheet[2]] == ["1001", 0.91]


def test_export_includes_semantic_score_for_default_columns_when_present(tmp_path: Path) -> None:
    student = StudentCurrent(
        STUD_ID="1001",
        STUD_NAME="Student",
        WSP_TECHNICAL_SKILLS="Python",
    )
    result = FilterResult(
        rows=(student,),
        selected_rows=(),
        total_count=1,
        page=1,
        page_size=50,
        applied_filter_count=1,
        applied_filter_metadata={},
        semantic_scores={"1001": 0.82},
    )

    export = export_filtered_results_to_excel(result, tmp_path)
    worksheet = load_workbook(export.path)["Filtered Results"]

    headers = [cell.value for cell in worksheet[1]]
    values = [cell.value for cell in worksheet[2]]
    assert "semantic_score" in headers
    assert values[headers.index("semantic_score")] == 0.82


def test_export_includes_semantic_explanation_for_default_columns_when_present(tmp_path: Path) -> None:
    student = StudentCurrent(STUD_ID="1001", STUD_NAME="Student")
    result = FilterResult(
        rows=(student,),
        selected_rows=(),
        total_count=1,
        page=1,
        page_size=50,
        applied_filter_count=1,
        applied_filter_metadata={},
        semantic_scores={"1001": 0.82},
        semantic_reasons={"1001": "Embedding match 0.82. Technical skills: Excel reporting."},
    )

    export = export_filtered_results_to_excel(result, tmp_path)
    worksheet = load_workbook(export.path)["Filtered Results"]

    headers = [cell.value for cell in worksheet[1]]
    values = [cell.value for cell in worksheet[2]]
    assert "semantic_explanation" in headers
    assert values[headers.index("semantic_explanation")] == "Embedding match 0.82. Technical skills: Excel reporting."


def metadata_rows(path: Path) -> dict[str, str]:
    worksheet = load_workbook(path)["Filter Metadata"]
    return {
        str(key): str(value)
        for key, value in worksheet.iter_rows(min_row=2, values_only=True)
    }


def test_export_metadata_sheet_exists_and_stores_filter_json(tmp_path: Path) -> None:
    result = FilterResult(
        rows=(),
        selected_rows=({"STUD_ID": "1001"},),
        total_count=1,
        page=1,
        page_size=50,
        applied_filter_count=1,
        applied_filter_metadata={"text_filters": [{"field_name": "STUD_NAME", "operator": "contains", "value": "Alice"}]},
    )

    export = export_filtered_results_to_excel(result, tmp_path)
    workbook = load_workbook(export.path)
    metadata = metadata_rows(export.path)

    assert "Filter Metadata" in workbook.sheetnames
    assert "Alice" in metadata["filter_json"]


def test_export_metadata_stores_timestamp_row_count_and_app_version(tmp_path: Path) -> None:
    exported_at = datetime(2026, 6, 4, 14, 1, 5, tzinfo=UTC)
    result = FilterResult(
        rows=(),
        selected_rows=({"STUD_ID": "1001"}, {"STUD_ID": "1002"}),
        total_count=2,
        page=1,
        page_size=50,
        applied_filter_count=0,
        applied_filter_metadata={},
    )

    export = export_filtered_results_to_excel(result, tmp_path, exported_at=exported_at, app_version="9.9.9")
    metadata = metadata_rows(export.path)

    assert metadata["export_timestamp"] == "2026-06-04T14:01:05+00:00"
    assert metadata["number_of_rows"] == "2"
    assert metadata["app_version"] == "9.9.9"


def test_export_metadata_stores_source_batch_ids(tmp_path: Path) -> None:
    engine = create_sqlite_engine(tmp_path / "wsp.db")
    initialize_database(engine)
    session_factory = create_session_factory(engine)

    with session_factory() as session:
        first_batch = ImportBatch(filename="one.xlsx", file_path="C:/one.xlsx", file_hash="one")
        second_batch = ImportBatch(filename="two.xlsx", file_path="C:/two.xlsx", file_hash="two")
        session.add_all([first_batch, second_batch])
        session.commit()
        session.add_all(
            [
                StudentCurrent(STUD_ID="1001", STUD_NAME="Alice", last_seen_batch_id=second_batch.batch_id),
                StudentCurrent(STUD_ID="1002", STUD_NAME="Bob", last_seen_batch_id=first_batch.batch_id),
            ]
        )
        session.commit()

    with session_factory() as session:
        filter_result = execute_filter_request(session, FilterRequest(selected_columns=("STUD_ID",)))
        export = export_filtered_results_to_excel(filter_result, tmp_path)

    metadata = metadata_rows(export.path)

    assert metadata["source_batch_ids"] == "1,2"


def test_export_log_row_is_created(tmp_path: Path) -> None:
    engine = create_sqlite_engine(tmp_path / "wsp.db")
    initialize_database(engine)
    session_factory = create_session_factory(engine)
    result = FilterResult(
        rows=(),
        selected_rows=({"STUD_ID": "1001"},),
        total_count=1,
        page=1,
        page_size=50,
        applied_filter_count=0,
        applied_filter_metadata={},
    )

    with session_factory() as session:
        filter_run = log_filter_run(session, request=FilterRequest(), result_count=1)
        export = export_filtered_results_to_excel(result, tmp_path, session=session, filter_run_id=filter_run.filter_run_id)
        session.commit()
        export_log_id = export.export_log_id

    with session_factory() as session:
        export_log = session.get(ExportLog, export_log_id)

    assert export_log is not None
    assert Path(export_log.export_path).exists()
    assert export_log.row_count == 1


def test_export_log_links_to_filter_run(tmp_path: Path) -> None:
    engine = create_sqlite_engine(tmp_path / "wsp.db")
    initialize_database(engine)
    session_factory = create_session_factory(engine)
    result = FilterResult(
        rows=(),
        selected_rows=({"STUD_ID": "1001"},),
        total_count=1,
        page=1,
        page_size=50,
        applied_filter_count=0,
        applied_filter_metadata={},
    )

    with session_factory() as session:
        filter_run = log_filter_run(session, request=FilterRequest(), result_count=1)
        export = export_filtered_results_to_excel(result, tmp_path, session=session, filter_run_id=filter_run.filter_run_id)
        session.commit()

    with session_factory() as session:
        export_log = session.get(ExportLog, export.export_log_id)

    assert export_log.filter_run_id == filter_run.filter_run_id


def test_export_logging_requires_session_and_filter_run_together(tmp_path: Path) -> None:
    result = FilterResult(
        rows=(),
        selected_rows=({"STUD_ID": "1001"},),
        total_count=1,
        page=1,
        page_size=50,
        applied_filter_count=0,
        applied_filter_metadata={},
    )

    with pytest.raises(ValueError, match="provided together"):
        export_filtered_results_to_excel(result, tmp_path, filter_run_id=1)
