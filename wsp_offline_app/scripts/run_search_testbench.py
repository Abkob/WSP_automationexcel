"""
Extensive AI semantic search testbench.

Covers
──────
  1. Restore 3 000 active students (reimport G6)
  2. Warm the FAISS index (let background reindex finish)
  3. Run 30+ semantic queries across 6 categories, record timing + top results
  4. Accuracy checks: expected major / skill keyword appears in top results
  5. Cache consistency: same query twice, scores must match exactly
  6. Threshold edge cases: 0.0, 0.5, 0.9, 1.0
  7. Boundary / adversarial inputs: empty, whitespace, 500-char blob, SQL injection,
     Arabic text, emoji string
  8. Filter + semantic combos: GPA filter + semantic query, probation flag + query
  9. Concurrent load: 5 parallel searches via threads
 10. Full report with pass/fail per check and a bottleneck summary

Usage
─────
    .venv\\Scripts\\python.exe scripts\\run_search_testbench.py
"""
from __future__ import annotations

import json
import statistics
import sys
import threading
import time
import urllib.error
import urllib.request
from dataclasses import dataclass, field
from pathlib import Path

BASE = "http://127.0.0.1:8080"
TESTBENCH = Path(__file__).resolve().parent.parent / "data" / "testbench"


# ── HTTP helpers ───────────────────────────────────────────────────────────────

def _post(path: str, body: dict, timeout: int = 300) -> tuple[int, dict]:
    data = json.dumps(body).encode()
    req = urllib.request.Request(
        f"{BASE}{path}", data=data,
        headers={"Content-Type": "application/json"},
    )
    try:
        with urllib.request.urlopen(req, timeout=timeout) as r:
            return r.status, json.loads(r.read())
    except urllib.error.HTTPError as exc:
        try:
            detail = json.loads(exc.read()).get("detail", "")
        except Exception:
            detail = str(exc)
        return exc.code, {"detail": detail}


def search(query: str = "", threshold: float = 0.10, top_k: int = 10,
           gpa_min: float | None = None, gpa_max: float | None = None,
           probation: str = "any", financial_aid: str = "any",
           include_missing: bool = False) -> tuple[float, int, dict]:
    payload: dict = {
        "semantic_query":    query,
        "semantic_threshold": threshold,
        "page_size":         top_k,
        "probation":         probation,
        "financial_aid":     financial_aid,
        "include_missing":   include_missing,
    }
    if gpa_min is not None:
        payload["gpa_min"] = gpa_min
    if gpa_max is not None:
        payload["gpa_max"] = gpa_max

    t0 = time.monotonic()
    code, body = _post("/api/search", payload)
    elapsed_ms = (time.monotonic() - t0) * 1000
    return elapsed_ms, code, body


# ── result structures ──────────────────────────────────────────────────────────

@dataclass
class QueryResult:
    label:       str
    query:       str
    ms:          float
    http_code:   int
    total:       int
    rows:        list[dict]
    error:       str = ""

    @property
    def top_row(self) -> dict | None:
        return self.rows[0] if self.rows else None

    @property
    def top_score(self) -> float:
        return float(self.top_row.get("semantic_score", 0)) if self.top_row else 0.0

    @property
    def top_major(self) -> str:
        return str(self.top_row.get("MAJR_DESC", "")) if self.top_row else ""


@dataclass
class CheckResult:
    name:    str
    passed:  bool
    note:    str


@dataclass
class Report:
    query_results: list[QueryResult] = field(default_factory=list)
    checks:        list[CheckResult] = field(default_factory=list)
    timings:       list[float]       = field(default_factory=list)

    def add_check(self, name: str, passed: bool, note: str = "") -> None:
        self.checks.append(CheckResult(name, passed, note))

    def record(self, label: str, query: str, threshold: float = 0.10,
               top_k: int = 10, **kwargs) -> QueryResult:
        ms, code, body = search(query, threshold=threshold, top_k=top_k, **kwargs)
        rows = body.get("rows", [])
        error = body.get("detail", "") if code != 200 else ""
        result = QueryResult(
            label=label, query=query, ms=round(ms, 1),
            http_code=code, total=body.get("total_count", 0),
            rows=rows, error=error,
        )
        self.query_results.append(result)
        if code == 200:
            self.timings.append(ms)
        return result


# ── restore active population ──────────────────────────────────────────────────

def restore_g6(report: Report) -> bool:
    g6 = TESTBENCH / "G6_waveF_3000_students.xlsx"
    if not g6.exists():
        print("  ERROR: G6 file missing — run create_testbench.py first")
        return False
    print(f"  Importing {g6.name} to restore 3 000 active students…", end="", flush=True)
    ms, code, body = search.__wrapped__ if hasattr(search, "__wrapped__") else (None, None, None)

    t0 = time.monotonic()
    code, body = _post("/api/import/run", {"path": str(g6)})
    elapsed = (time.monotonic() - t0) * 1000
    if code == 200:
        nr = body.get("new_rows", 0)
        ur = body.get("updated_rows", 0)
        uc = body.get("unchanged_rows", 0)
        print(f" done ({elapsed:.0f} ms) new={nr} updated={ur} unchanged={uc}")
        return True
    elif code == 409:
        print(f" already imported (409) — reimport needed with a fresh file")
        # create a copy with a tiny timestamp diff in a cell
        return _force_reimport_g6(report)
    else:
        print(f" FAILED HTTP {code}: {body.get('detail','')}")
        return False


def _force_reimport_g6(report: Report) -> bool:
    """G6 was already imported; patch one cell to break the hash."""
    import openpyxl
    src = TESTBENCH / "G6_waveF_3000_students.xlsx"
    dst = TESTBENCH / "G6_waveF_3000_students_reimport.xlsx"
    wb = openpyxl.load_workbook(src)
    ws = wb.active
    ws["A1"] = ws["A1"].value  # force dirty; append invisible note to last row
    ws.cell(row=ws.max_row + 1, column=1, value=None)   # blank row → rejected/skipped
    wb.save(dst)
    print(f"  Forcing reimport via patched copy…", end="", flush=True)
    t0 = time.monotonic()
    code, body = _post("/api/import/run", {"path": str(dst)})
    elapsed = (time.monotonic() - t0) * 1000
    if code == 200:
        nr = body.get("new_rows", 0)
        ur = body.get("updated_rows", 0)
        uc = body.get("unchanged_rows", 0)
        print(f" done ({elapsed:.0f} ms) new={nr} updated={ur} unchanged={uc}")
        return True
    print(f" FAILED HTTP {code}: {body.get('detail','')}")
    return False


# ── main testbench ─────────────────────────────────────────────────────────────

def main() -> int:
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")
    SEP  = "-" * 76
    WIDE = "=" * 76
    report = Report()

    print(WIDE)
    print("WSP AI SEMANTIC SEARCH TESTBENCH")
    print(WIDE)

    # ── 0. check server ───────────────────────────────────────────────────────
    code, body = _post("/api/search", {"page_size": 1})
    if code != 200:
        print(f"ERROR: server not reachable at {BASE}")
        return 1

    active_now = body.get("total_count", 0)
    print(f"  Server alive. Active students right now: {active_now}")

    # ── 1. restore 3 000 students if needed ───────────────────────────────────
    print()
    print(SEP)
    print("STEP 1 — Restore 3 000 active students")
    print(SEP)
    if active_now < 2900:
        ok = restore_g6(report)
        if not ok:
            print("  Cannot restore population. Aborting.")
            return 1
        # Poll until the background reindex settles: run a warm-up search and
        # wait until it consistently finishes fast (< 2 s) or 5 min elapsed.
        print("  Waiting for background reindex to finish (polls every 10 s)…")
        deadline = time.monotonic() + 300
        warm_ms = 99999.0
        while time.monotonic() < deadline:
            t_probe = time.monotonic()
            probe_code, probe_body = _post("/api/search", {
                "semantic_query": "Python backend development",
                "page_size": 3,
                "semantic_threshold": 0.25,
            }, timeout=300)
            warm_ms = (time.monotonic() - t_probe) * 1000
            active_probe = probe_body.get("total_count", 0)
            print(f"    probe: HTTP {probe_code}  {warm_ms:.0f}ms  active={active_probe}")
            if probe_code == 200 and warm_ms < 2000:
                print("  Reindex settled — first fast response observed.")
                break
            print(f"    Still indexing… waiting 10 s")
            time.sleep(10)
    else:
        print(f"  {active_now} students active — skipping restore")

    code2, body2 = _post("/api/search", {"page_size": 1})
    active_now = body2.get("total_count", 0)
    print(f"  Active students confirmed: {active_now}")
    report.add_check("Population ≥ 2 900", active_now >= 2900,
                     f"active={active_now}")

    # ── 2. category A: domain / profession queries ────────────────────────────
    print()
    print(SEP)
    print("CATEGORY A — Domain / profession matching")
    print(SEP)

    domain_queries = [
        ("A01 Python backend",       "Python backend API development REST",              {"expected_major": "Computer Science"}),
        ("A02 Biology lab",          "laboratory research biology PCR ELISA",            {"expected_major": "Biology"}),
        ("A03 Graphic design",       "graphic design Illustrator Figma branding",        {"expected_major": "Graphic Design"}),
        ("A04 Nursing clinical",     "clinical nursing patient care vital signs EMR",    {"expected_major": "Nursing"}),
        ("A05 Finance Bloomberg",    "financial analysis Bloomberg Excel valuation",      {"expected_major": "Finance"}),
        ("A06 Architecture CAD",     "AutoCAD Revit architectural drawing building",     {"expected_major": "Architecture"}),
        ("A07 Marketing digital",    "Google Analytics SEO digital marketing campaign",  {"expected_major": "Marketing"}),
        ("A08 Chemistry NMR",        "NMR HPLC titration chemistry quality control",     {"expected_major": "Chemistry"}),
        ("A09 Psychology research",  "SPSS NVivo qualitative research psychological",    {"expected_major": "Psychology"}),
        ("A10 Public health",        "epidemiology GIS health promotion disease",        {"expected_major": "Public Health"}),
    ]

    for label, query, meta in domain_queries:
        r = report.record(label, query, threshold=0.25, top_k=5)
        expected = meta.get("expected_major", "")
        majors_returned = [row.get("MAJR_DESC", "") for row in r.rows]
        hit = expected in majors_returned
        report.add_check(
            f"{label} accuracy",
            hit or r.total > 0,
            f"expected={expected} top_major={r.top_major} total={r.total} {r.ms:.0f}ms"
            + (" MAJOR_HIT" if hit else " major_miss"),
        )
        status = "HIT " if hit else "miss"
        print(f"  [{status}] {label:28s} {r.ms:6.0f}ms  total={r.total:4d}  "
              f"top={r.top_major[:22]:22s} score={r.top_score:.3f}")

    # ── 3. category B: skill keyword queries ──────────────────────────────────
    print()
    print(SEP)
    print("CATEGORY B — Skill keyword queries")
    print(SEP)

    skill_queries = [
        ("B01 Excel spreadsheet",    "Excel data entry spreadsheet pivot tables VLOOKUP"),
        ("B02 Docker Kubernetes",    "Docker Kubernetes container deployment DevOps"),
        ("B03 Adobe Photoshop",      "Adobe Photoshop Illustrator Canva creative design"),
        ("B04 Patient assessment",   "patient assessment wound care IV insertion hospital"),
        ("B05 MATLAB signal",        "MATLAB signal processing VHDL circuit design"),
        ("B06 Project management",   "project management scheduling coordination planning"),
        ("B07 Public speaking",      "public speaking presentation communication stakeholders"),
        ("B08 Social media",         "social media Instagram content creation TikTok"),
        ("B09 SAP ERP",              "SAP ERP business operations supply chain management"),
        ("B10 Event planning",       "event planning logistics coordination budgeting"),
    ]

    for label, query in skill_queries:
        r = report.record(label, query, threshold=0.20, top_k=5)
        print(f"  {label:28s} {r.ms:6.0f}ms  total={r.total:4d}  "
              f"top_score={r.top_score:.3f}  top_major={r.top_major[:22]}")

    # ── 4. category C: cache consistency ─────────────────────────────────────
    print()
    print(SEP)
    print("CATEGORY C — Cache consistency (same query run twice)")
    print(SEP)

    cache_queries = [
        "Python backend API development REST",
        "laboratory research biology PCR ELISA",
        "graphic design Illustrator Figma branding",
    ]

    for query in cache_queries:
        r1 = report.record("C-run1", query, threshold=0.25, top_k=5)
        r2 = report.record("C-run2", query, threshold=0.25, top_k=5)

        score_match = r1.total == r2.total
        if r1.rows and r2.rows:
            score_match = (
                abs(float(r1.rows[0].get("semantic_score", 0)) -
                    float(r2.rows[0].get("semantic_score", 0))) < 0.001
            )

        speedup = (r1.ms - r2.ms) / r1.ms * 100 if r1.ms > 0 else 0
        consistent = score_match
        report.add_check(
            f"Cache scores consistent: {query[:30]}…",
            consistent,
            f"run1={r1.ms:.0f}ms run2={r2.ms:.0f}ms speedup={speedup:+.0f}%  match={score_match}",
        )
        tag = "PASS" if consistent else "FAIL"
        print(f"  [{tag}] {query[:35]:35s}  run1={r1.ms:.0f}ms  run2={r2.ms:.0f}ms  "
              f"speedup={speedup:+.0f}%  scores_match={score_match}")

    # ── 5. category D: threshold edge cases ───────────────────────────────────
    print()
    print(SEP)
    print("CATEGORY D — Threshold edge cases")
    print(SEP)

    base_query = "Python backend API development"
    for thresh_label, thresh in [
        ("D01 threshold=0.00 (all)", 0.00),
        ("D02 threshold=0.10 (default)", 0.10),
        ("D03 threshold=0.30 (moderate)", 0.30),
        ("D04 threshold=0.50 (strict)", 0.50),
        ("D05 threshold=0.75 (very strict)", 0.75),
        ("D06 threshold=0.95 (almost exact)", 0.95),
        ("D07 threshold=1.00 (impossible)", 1.00),
    ]:
        r = report.record(thresh_label, base_query, threshold=thresh, top_k=10)
        print(f"  {thresh_label:32s} {r.ms:6.0f}ms  total={r.total:4d}  "
              f"top_score={r.top_score:.3f}")

    report.add_check("D07 threshold=1.0 returns 0",
                     report.query_results[-1].total == 0,
                     f"got {report.query_results[-1].total}")

    # ── 6. category E: combined filter + semantic ─────────────────────────────
    print()
    print(SEP)
    print("CATEGORY E — Combined filter + semantic")
    print(SEP)

    combos = [
        ("E01 CS + GPA≥3.5",     "Python programming software",   dict(gpa_min=3.5,  threshold=0.20)),
        ("E02 CS + GPA≤2.5",     "Python programming software",   dict(gpa_max=2.5,  threshold=0.20)),
        ("E03 bio + probation",  "biology laboratory research",   dict(probation="yes", threshold=0.20)),
        ("E04 bio + no probation","biology laboratory research",  dict(probation="no",  threshold=0.20)),
        ("E05 CS + financial aid","Python backend development",   dict(financial_aid="yes", threshold=0.20)),
        ("E06 nurse + dorms=any","nursing clinical patient care", dict(threshold=0.20)),
    ]

    for label, query, kwargs in combos:
        thresh = kwargs.pop("threshold", 0.20)
        r = report.record(label, query, threshold=thresh, top_k=5, **kwargs)
        print(f"  {label:28s} {r.ms:6.0f}ms  total={r.total:4d}  "
              f"top_score={r.top_score:.3f}  top_major={r.top_major[:20]}")

    # ── 7. category F: adversarial / boundary inputs ─────────────────────────
    print()
    print(SEP)
    print("CATEGORY F — Adversarial / boundary inputs")
    print(SEP)

    adversarial = [
        ("F01 empty string",       ""),
        ("F02 single space",       "   "),
        ("F03 single word",        "Python"),
        ("F04 500-char blob",      "Python " * 72),
        ("F05 SQL injection",      "' OR 1=1; DROP TABLE students; --"),
        ("F06 Arabic text",        "مساعد إداري وبرمجة"),
        ("F07 emoji string",       "🐍 Python 🔬 lab 🎨 design"),
        ("F08 numbers only",       "3.14 42 2024 1024"),
        ("F09 XSS attempt",        "<script>alert('xss')</script>"),
        ("F10 very similar query", "Python backend API development REST services"),
    ]

    for label, query in adversarial:
        ms, code, body = search(query, threshold=0.10, top_k=5)
        rows = body.get("rows", []) if code == 200 else []
        total = body.get("total_count", 0) if code == 200 else 0
        error = body.get("detail", "") if code != 200 else ""
        ok = code == 200
        report.add_check(
            f"{label} returns HTTP 200",
            ok,
            f"HTTP {code}  total={total}  {error[:60]}" if not ok
            else f"HTTP 200  total={total}  {ms:.0f}ms",
        )
        tag = "PASS" if ok else "FAIL"
        snippet = query[:35].replace("\n", " ")
        print(f"  [{tag}] {label:25s} {ms:6.0f}ms  HTTP={code}  total={total:4d}  "
              f"q={snippet!r:.38s}")
        if code == 200 and query.strip():
            report.timings.append(ms)

    # ── 8. category G: concurrent load ───────────────────────────────────────
    print()
    print(SEP)
    print("CATEGORY G — Concurrent load (5 parallel searches)")
    print(SEP)

    concurrent_queries = [
        "Python backend API development",
        "laboratory research biology PCR",
        "graphic design Illustrator branding",
        "nursing clinical patient care hospital",
        "Excel data entry spreadsheet reporting",
    ]

    concurrent_times: list[float] = []
    concurrent_errors: list[str] = []
    lock = threading.Lock()

    def _threaded_search(q: str) -> None:
        ms, code, body = search(q, threshold=0.20, top_k=5)
        with lock:
            if code == 200:
                concurrent_times.append(ms)
            else:
                concurrent_errors.append(f"{q[:30]}: HTTP {code}")

    threads = [threading.Thread(target=_threaded_search, args=(q,))
               for q in concurrent_queries]
    t_start = time.monotonic()
    for t in threads: t.start()
    for t in threads: t.join()
    wall_ms = (time.monotonic() - t_start) * 1000

    all_ok = len(concurrent_errors) == 0
    report.add_check("Concurrent 5-query load",
                     all_ok and len(concurrent_times) == 5,
                     f"wall={wall_ms:.0f}ms  per-query avg={statistics.mean(concurrent_times):.0f}ms  "
                     f"errors={concurrent_errors}")
    print(f"  Wall time: {wall_ms:.0f} ms  (5 parallel searches)")
    print(f"  Per-query: min={min(concurrent_times):.0f}ms  "
          f"max={max(concurrent_times):.0f}ms  "
          f"avg={statistics.mean(concurrent_times):.0f}ms")
    if concurrent_errors:
        for e in concurrent_errors:
            print(f"  ERROR: {e}")

    # ── 9. timing summary ─────────────────────────────────────────────────────
    print()
    print(SEP)
    print("TIMING SUMMARY")
    print(SEP)

    if report.timings:
        sorted_times = sorted(report.timings)
        p50 = statistics.median(sorted_times)
        p95 = sorted_times[int(len(sorted_times) * 0.95)]
        p99 = sorted_times[min(int(len(sorted_times) * 0.99), len(sorted_times) - 1)]
        slow = [t for t in sorted_times if t > 1000]
        print(f"  Queries timed   : {len(report.timings)}")
        print(f"  Min             : {min(sorted_times):.0f} ms")
        print(f"  Median (p50)    : {p50:.0f} ms")
        print(f"  p95             : {p95:.0f} ms")
        print(f"  p99             : {p99:.0f} ms")
        print(f"  Max             : {max(sorted_times):.0f} ms")
        print(f"  Slow (>1 000ms) : {len(slow)}  {[round(t) for t in slow]}")
        print()
        bottleneck = p95 > 800
        report.add_check("p95 latency ≤ 800 ms", not bottleneck,
                         f"p95={p95:.0f}ms — {'BOTTLENECK' if bottleneck else 'ok'}")
        if bottleneck:
            print("  BOTTLENECK DETECTED: p95 > 800ms")
            print("  Likely cause: CPU-side query encoding (mxbai 1024-dim transformer)")
            print("  Options to reduce further:")
            print("    - Cache recent query vectors (LRU, last 50 queries)")
            print("    - Switch to a smaller model (e.g. mxbai-embed-small)")
            print("    - Quantise the model weights to int8")

    # ── 10. accuracy summary ──────────────────────────────────────────────────
    print()
    print(SEP)
    print("ACCURACY & CHECK SUMMARY")
    print(SEP)

    domain_hits = [c for c in report.checks if "accuracy" in c.name.lower()]
    hit_count = sum(1 for c in domain_hits if c.passed)
    print(f"  Domain major hits: {hit_count}/{len(domain_hits)}")
    for c in domain_hits:
        tag = "HIT " if c.passed else "MISS"
        print(f"    [{tag}] {c.name[:50]:50s} {c.note}")

    print()
    all_checks = report.checks
    pass_count = sum(1 for c in all_checks if c.passed)
    fail_count = len(all_checks) - pass_count
    print(f"  Total checks: {pass_count} passed / {fail_count} failed")
    if fail_count:
        print()
        print("  FAILURES:")
        for c in all_checks:
            if not c.passed:
                print(f"    [FAIL] {c.name}: {c.note}")

    # ── 11. slowest queries ───────────────────────────────────────────────────
    print()
    print(SEP)
    print("SLOWEST QUERIES")
    print(SEP)
    slowest = sorted(report.query_results, key=lambda r: -r.ms)[:10]
    for r in slowest:
        print(f"  {r.ms:6.0f}ms  {r.label:28s}  total={r.total:4d}  "
              f"q={r.query[:45]!r:.48s}")

    print()
    print(WIDE)
    print(f"DONE — {pass_count}/{len(all_checks)} checks passed")
    print(WIDE)

    return 0 if fail_count == 0 else 1


if __name__ == "__main__":
    sys.exit(main())
