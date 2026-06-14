"""
Extensive import testbench — generates data/testbench/ with one XLSX per import scenario.

Groups
──────
  A  Core pipeline          sequential waves testing new/update/missing/restore/duplicate
  B  ID edge cases          integer IDs, float IDs, leading zeros, missing/blank IDs
  C  Field value edge cases boolean variants, malformed GPA, unicode, long text, whitespace
  D  Column schema variants extra columns, minimal columns, partial columns, lowercase headers
  E  Sheet/workbook shape   multi-sheet, single-row, empty sheet, blank rows interspersed
  F  Filter & search targets designed data to exercise every filter control in the UI

Usage
─────
    .venv\\Scripts\\python.exe scripts\\create_testbench.py

Output
──────
    data/testbench/  (one .xlsx per scenario)
    A manifest is printed at the end with: filename, expected outcome, what to verify.
"""
from __future__ import annotations

import copy
import random
import shutil
from pathlib import Path

import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

# ── canonical column list ──────────────────────────────────────────────────────
COLS = [
    "STUD_ID", "STUD_NAME", "MAJR_DESC", "CLAS_DESC", "STUD_EMAIL",
    "WSP_WRITTEN_LANGUAGES", "WSP_SPOKEN_LANGUAGES",
    "WSP_ORGANIZATIONAL_SKILLS", "WSP_TECHNICAL_SKILLS",
    "WSP_INTERPERSONAL_SKILLS", "WSP_ADDITIONAL_SKILLS",
    "WSP_PREV_WORK", "WSP_PREVIOUS_TYPE_OF_WORK", "WSP_PREFERRED_TYPE_OF_WORK",
    "DEANS_WARNING", "ENRL_TERM", "DEAN_WARN", "MOBILE_NBR", "PROBATION",
    "APPLICATION_DATE", "CUM_GPA", "STST_DESC", "STYP_CODE", "STYP_DESC",
    "ENROLLED_IND", "REGISTERED_IND", "LEVL_CODE", "COLL_CODE",
    "TOTAL_CREDIT_HOURS", "ASTD_TERM", "ATSD_CODE_END_OF_TERM", "ASTD_DESC",
    "ASTD_DATE_END_OF_TERM", "USAID", "MASTER_CARD", "UPP_MEPI", "GAS",
    "FINANCIAL_AID", "DORMS",
]

# ── data pools ─────────────────────────────────────────────────────────────────
_RNG = random.Random(99)

_MAJORS = [
    ("Computer Science",       "EN",  "Python, SQL, Git",       "Software development"),
    ("Electrical Engineering", "EN",  "MATLAB, VHDL, C",        "Hardware design"),
    ("Business Administration","OSB", "Excel, PowerPoint, SAP", "Operations management"),
    ("Finance",                "OSB", "Bloomberg, Excel, Python","Financial analysis"),
    ("Biology",                "AS",  "PCR, ELISA, microscopy", "Lab assistance"),
    ("Chemistry",              "AS",  "NMR, HPLC, titration",   "Quality control"),
    ("Psychology",             "AS",  "SPSS, NVivo, Excel",     "Research assistance"),
    ("Architecture",           "MSFEA","AutoCAD, Revit, SketchUp","Architectural drawing"),
    ("Nursing",                "HS",  "Patient assessment, EMR","Clinical support"),
    ("Public Health",          "HS",  "SPSS, epidemiology, GIS","Health promotion"),
    ("English Literature",     "AH",  "Academic writing, editing","Content editing"),
    ("Political Science",      "AH",  "ArcGIS, policy analysis","Policy research"),
    ("Graphic Design",         "MSFEA","Adobe Illustrator, Figma","Visual content design"),
    ("Marketing",              "OSB", "Google Analytics, SEO, Canva","Digital marketing"),
    ("Accounting",             "OSB", "QuickBooks, Excel, IFRS","Audit support"),
]
_CLASSES = ["Freshman", "Sophomore", "Junior", "Senior"]
_FIRST   = ["Ahmad", "Lara", "Ali", "Maya", "Omar", "Sara", "Jad", "Nour",
             "Rami", "Dana", "Tarek", "Hana", "Fadi", "Rima", "Elie", "Mona",
             "Georges", "Nadine", "Karim", "Aya", "Bilal", "Celine", "Wissam",
             "Rita", "Ramzi", "Carla", "Khaled", "Maria", "Nader", "Lynn",
             "Samir", "Joelle", "Ghassan", "Sandra", "Tony", "Jana", "Ziad",
             "Rania", "Rabih", "Layla", "Patrick", "Dina", "Roland", "Ghina",
             "Simon", "Roula", "Nabil", "Marianne", "Walid", "Christelle"]
_FAMILY  = ["Hassan", "Khalil", "Nassar", "Daher", "Frem", "Moussa", "Rizk",
             "Khoury", "Salameh", "Habib", "Makhoul", "Haddad", "Sfeir", "Azar",
             "Hamdan", "Saad", "Hajj", "Kanaan", "Farhat", "Zein", "Raad",
             "Jaber", "Mansour", "Kassis", "Gemayel", "Karam", "Abdallah",
             "Francis", "Maalouf", "Ghanem", "Chaaban", "Turk", "Sawaya",
             "Lahoud", "Abboud", "Bou Khalil", "Nakad", "Tyan", "Abi Nader", "Frem"]
_ORG = ["Time management, scheduling", "Event planning, coordination",
        "Project management, reporting", "Record keeping, documentation",
        "Research organisation, note taking", "Budget tracking, logistics"]
_INTER = ["Communication, teamwork", "Critical thinking, problem solving",
          "Presentation skills, stakeholder engagement",
          "Attention to detail, patience", "Collaboration, conflict resolution"]
_ADD   = ["Microsoft Office, Canva", "Git, Linux basics", "Public speaking",
          "Social media management", "Data entry, Excel formulas",
          "Google Workspace, Notion", "First aid certified"]


def _make_student(sid: int, *, prob: bool = False, aid: bool = False,
                   dorms: bool = False, major_idx: int | None = None,
                   cls: str | None = None, rng: random.Random = _RNG) -> dict:
    fname = rng.choice(_FIRST)
    lname = rng.choice(_FAMILY)
    if major_idx is None:
        major_idx = rng.randrange(len(_MAJORS))
    majr, coll, skills, pref = _MAJORS[major_idx % len(_MAJORS)]
    cls = cls or rng.choice(_CLASSES)
    credits = {"Freshman": rng.randint(0,30), "Sophomore": rng.randint(31,60),
               "Junior": rng.randint(61,90), "Senior": rng.randint(91,130)}[cls]
    gpa = round(rng.uniform(1.50, 2.29) if prob else rng.uniform(2.30, 4.00), 2)
    astd_c = "AP" if gpa < 2.0 else ("DL" if gpa >= 3.5 else "GS")
    astd_d = {"AP": "Academic Probation", "DL": "Dean's List", "GS": "Good Standing"}[astd_c]
    warn = "Y" if prob else "N"
    sid_str = str(200000 + sid).zfill(6)
    email = f"{fname[0].lower()}{lname[:4].lower()}{sid:03d}@mail.aub.edu"
    prev_options = [
        f"Teaching assistant — {majr.split()[0]} department",
        f"Research assistant — {majr.split()[0]} lab",
        "Intern — local startup",
        "None",
    ]
    prev = rng.choice(prev_options)
    prev_type = ("Academic support" if "assistant" in prev else
                 "Internship" if "Intern" in prev else "None")
    return {
        "STUD_ID": sid_str,
        "STUD_NAME": f"{fname} {lname}",
        "MAJR_DESC": majr,
        "CLAS_DESC": cls,
        "STUD_EMAIL": email,
        "WSP_WRITTEN_LANGUAGES": "Arabic, English",
        "WSP_SPOKEN_LANGUAGES": "Arabic, English",
        "WSP_ORGANIZATIONAL_SKILLS": rng.choice(_ORG),
        "WSP_TECHNICAL_SKILLS": skills,
        "WSP_INTERPERSONAL_SKILLS": rng.choice(_INTER),
        "WSP_ADDITIONAL_SKILLS": rng.choice(_ADD),
        "WSP_PREV_WORK": prev,
        "WSP_PREVIOUS_TYPE_OF_WORK": prev_type,
        "WSP_PREFERRED_TYPE_OF_WORK": pref,
        "DEANS_WARNING": warn,
        "ENRL_TERM": "202520",
        "DEAN_WARN": warn,
        "MOBILE_NBR": f"+9617{rng.randint(1000000,9999999)}",
        "PROBATION": "Y" if prob else "N",
        "APPLICATION_DATE": "2025-09-01",
        "CUM_GPA": gpa,
        "STST_DESC": "Active",
        "STYP_CODE": "R",
        "STYP_DESC": "Regular",
        "ENROLLED_IND": "Y",
        "REGISTERED_IND": "Y",
        "LEVL_CODE": "UG",
        "COLL_CODE": coll,
        "TOTAL_CREDIT_HOURS": credits,
        "ASTD_TERM": "202420",
        "ATSD_CODE_END_OF_TERM": astd_c,
        "ASTD_DESC": astd_d,
        "ASTD_DATE_END_OF_TERM": "2024-06-01",
        "USAID": "N",
        "MASTER_CARD": "N",
        "UPP_MEPI": "N",
        "GAS": "N",
        "FINANCIAL_AID": "Y" if aid else "N",
        "DORMS": "Y" if dorms else "N",
    }


def _write_wb(rows: list[dict], path: Path, columns: list[str] = COLS,
              sheet_name: str = "Sheet1",
              data_keys: list[str] | None = None) -> None:
    """data_keys: if supplied, use these keys to read row values (positionally
    aligned with columns). Allows writing alternate headers while still reading
    data from dicts with canonical UPPERCASE keys."""
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="6B1A2A")  # AUB maroon

    for col_idx, col in enumerate(columns, 1):
        cell = ws.cell(row=1, column=col_idx, value=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    keys = data_keys if data_keys is not None else columns
    for row_idx, row in enumerate(rows, 2):
        for col_idx, key in enumerate(keys, 1):
            ws.cell(row=row_idx, column=col_idx, value=row.get(key))

    # auto-width (approximate)
    for col_cells in ws.columns:
        length = max((len(str(c.value or "")) for c in col_cells), default=8)
        ws.column_dimensions[col_cells[0].column_letter].width = min(length + 2, 40)

    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)


def _students(start: int, count: int, **kwargs) -> list[dict]:
    return [_make_student(start + i, **kwargs) for i in range(count)]


# ══════════════════════════════════════════════════════════════════════════════
# GROUP A — Core import pipeline (sequential waves)
# ══════════════════════════════════════════════════════════════════════════════

def gen_A(out: Path) -> list[dict]:
    manifest = []

    # ── A1: clean first import ───────────────────────────────────────────────
    a1 = _students(1, 30)
    _write_wb(a1, out / "A1_first_import_30.xlsx")
    manifest.append(("A1_first_import_30.xlsx",
        "IMPORT FIRST (clean DB)",
        "new=30  updated=0  unchanged=0  missing=0",
        "Dashboard shows 30 active students. No history rows."))

    # ── A2: 30 unchanged + 10 new + 5 updates ───────────────────────────────
    a2 = copy.deepcopy(a1)
    changed_ids = []
    for i in [0, 4, 9, 14, 19]:                     # positions to edit
        a2[i]["CUM_GPA"] = round(a2[i]["CUM_GPA"] + 0.25, 2)
        a2[i]["WSP_TECHNICAL_SKILLS"] += ", Docker"
        changed_ids.append(a2[i]["STUD_ID"])
    a2 += _students(31, 10)
    _write_wb(a2, out / "A2_add_update.xlsx")
    manifest.append(("A2_add_update.xlsx",
        "IMPORT SECOND (after A1)",
        "new=10  updated=5  unchanged=25  missing=0",
        f"5 students gain GPA bump + Docker skill. History rows created for: {', '.join(changed_ids[:3])}…"))

    # ── A3: drop 8 → go missing, 35 unchanged, 2 new ────────────────────────
    a2_all_ids = {s["STUD_ID"]: s for s in a2}
    drop_ids = {a1[i]["STUD_ID"] for i in [2, 5, 11, 17, 22, 25, 28, 29]}
    a3 = [copy.deepcopy(a2_all_ids[s["STUD_ID"]]) for s in a2 if s["STUD_ID"] not in drop_ids]
    a3 += _students(45, 2)
    _write_wb(a3, out / "A3_go_missing.xlsx")
    manifest.append(("A3_go_missing.xlsx",
        "IMPORT THIRD (after A2)  — 8 students absent",
        "new=2  updated=0  unchanged=32  missing=8",
        f"8 students flagged missing_from_latest_import=True. IDs: {', '.join(sorted(drop_ids)[:4])}…"))

    # ── A4: restore missing with SAME data → restored_student ───────────────
    a4_all = {s["STUD_ID"]: s for s in a2}          # original data from A2
    a4 = [copy.deepcopy(a4_all[s["STUD_ID"]]) for s in a2]  # bring everyone back
    a4 += _students(47, 1)
    _write_wb(a4, out / "A4_restore_unchanged.xlsx")
    manifest.append(("A4_restore_unchanged.xlsx",
        "IMPORT FOURTH — restores 8 missing students with IDENTICAL data",
        "new=1  updated=0  unchanged=32  restored=8  missing=0",
        "8 restored_student events in history (hash unchanged). unchanged_rows count rises."))

    # ── A5: restore missing with CHANGED data → updated_student ─────────────
    a5_base = {s["STUD_ID"]: copy.deepcopy(s) for s in a2}
    for sid in list(drop_ids)[:4]:                   # change 4 of the 8 restored
        a5_base[sid]["CUM_GPA"] = round(a5_base[sid]["CUM_GPA"] - 0.30, 2)
        a5_base[sid]["WSP_PREFERRED_TYPE_OF_WORK"] += " (updated on return)"
    a5 = list(a5_base.values()) + _students(48, 1)
    _write_wb(a5, out / "A5_restore_updated.xlsx")
    manifest.append(("A5_restore_updated.xlsx",
        "IMPORT FIFTH — restores missing students; 4 return with changed data",
        "new=1  updated=4  unchanged=36  missing=0",
        "4 of the restored students have changed GPA/prefs → updated_student (not restored_student)."))

    # ── A6: duplicate of A1 → rejected ──────────────────────────────────────
    a6_path = out / "A6_DUPLICATE_of_A1.xlsx"
    shutil.copy2(out / "A1_first_import_30.xlsx", a6_path)
    manifest.append(("A6_DUPLICATE_of_A1.xlsx",
        "IMPORT — exact byte copy of A1  ← SHOULD BE REJECTED",
        "Expected: DuplicateExcelFileError — import aborted, nothing written",
        "Import page should show an error banner. DB remains unchanged."))

    return manifest


# ══════════════════════════════════════════════════════════════════════════════
# GROUP B — STUD_ID edge cases
# ══════════════════════════════════════════════════════════════════════════════

def gen_B(out: Path) -> list[dict]:
    manifest = []

    # ── B1: STUD_ID stored as Excel integer (no quotes) ──────────────────────
    base = _students(100, 5)
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws.append(COLS)
    for s in base:
        row = [s.get(c) for c in COLS]
        row[0] = int(s["STUD_ID"])       # integer, not string
        ws.append(row)
    (out / "B1_integer_ids.xlsx").parent.mkdir(parents=True, exist_ok=True)
    wb.save(out / "B1_integer_ids.xlsx")
    manifest.append(("B1_integer_ids.xlsx",
        "IMPORT — STUD_ID column stored as integer in Excel",
        "new=5  (IDs normalised from int → string by importer)",
        "Students imported correctly. IDs stored as strings in DB."))

    # ── B2: STUD_ID stored as float (200001.0) ───────────────────────────────
    wb2 = Workbook()
    ws2 = wb2.active
    ws2.title = "Sheet1"
    ws2.append(COLS)
    for s in base:
        row = [s.get(c) for c in COLS]
        row[0] = float(s["STUD_ID"])    # 200001.0
        ws2.append(row)
    wb2.save(out / "B2_float_ids.xlsx")
    manifest.append(("B2_float_ids.xlsx",
        "IMPORT — STUD_ID stored as float (200001.0) in Excel",
        "new=5  (importer strips .0 suffix)",
        "Students imported correctly. IDs '200001' not '200001.0'."))

    # ── B3: leading-zero IDs (060001, 060002 …) ──────────────────────────────
    lo_students = []
    for i in range(5):
        s = copy.deepcopy(base[i])
        s["STUD_ID"] = f"0{60000 + i + 1}"   # 060001 … 060005
        lo_students.append(s)
    _write_wb(lo_students, out / "B3_leading_zero_ids.xlsx")
    manifest.append(("B3_leading_zero_ids.xlsx",
        "IMPORT — STUD_IDs with leading zeros (060001 … 060005)",
        "new=5  (leading zeros must be preserved)",
        "IDs in DB must be '060001' not '60001'. Check Excel Sheets view."))

    # ── B4: some rows have blank/null STUD_ID → those rows rejected ──────────
    mix = _students(200, 8)
    bad_rows = [
        {c: (None if c == "STUD_ID" else mix[0].get(c)) for c in COLS},   # null ID
        {c: ("   " if c == "STUD_ID" else mix[1].get(c)) for c in COLS},  # blank spaces
        {c: ("" if c == "STUD_ID" else mix[2].get(c)) for c in COLS},     # empty string
    ]
    mixed = mix[3:] + bad_rows          # 5 valid + 3 bad (interleaved at end)
    _write_wb(mixed, out / "B4_missing_id_rows.xlsx")
    manifest.append(("B4_missing_id_rows.xlsx",
        "IMPORT — 5 valid rows + 3 rows with null/blank STUD_ID",
        "new=5  row_rejected=3  (3 rows logged and skipped)",
        "Import completes. 3 rejection events in import log. Check Import Pipeline section."))

    return manifest


# ══════════════════════════════════════════════════════════════════════════════
# GROUP C — Field value edge cases
# ══════════════════════════════════════════════════════════════════════════════

def gen_C(out: Path) -> list[dict]:
    manifest = []

    # ── C1: boolean field variants ────────────────────────────────────────────
    bool_variants = [
        ("Y", "N", "Y"),   # standard
        ("N", "Y", "N"),   # standard inverse
        ("TRUE", "FALSE", "TRUE"),    # Excel TRUE/FALSE
        ("true", "false", "true"),    # lowercase
        ("Yes", "No", "Yes"),         # human-readable
        ("yes", "no", "yes"),         # lowercase human
        ("1", "0", "1"),              # numeric
    ]
    c1_rows = []
    for i, (prob, aid, dorms) in enumerate(bool_variants):
        s = _make_student(300 + i)
        s["PROBATION"] = prob
        s["FINANCIAL_AID"] = aid
        s["DORMS"] = dorms
        s["DEANS_WARNING"] = prob
        s["DEAN_WARN"] = prob
        s["ENROLLED_IND"] = "Y"
        s["REGISTERED_IND"] = "Y"
        s["USAID"] = "N"
        s["MASTER_CARD"] = "N"
        s["UPP_MEPI"] = "N"
        s["GAS"] = "N"
        c1_rows.append(s)
    _write_wb(c1_rows, out / "C1_boolean_variants.xlsx")
    manifest.append(("C1_boolean_variants.xlsx",
        "IMPORT — boolean fields as Y/N, TRUE/FALSE, 1/0, Yes/No",
        "new=7  (all rows imported; boolean values stored as-is)",
        "All 7 students import. Check PROBATION/FINANCIAL_AID/DORMS values in Excel Sheets view."))

    # ── C2: malformed CUM_GPA values ─────────────────────────────────────────
    c2_rows = []
    gpa_cases = [3.75, 2.00, 0.00, 4.00, None, "N/A", "3.5x", "n/a", 5.00, -1.00, ""]
    for i, gpa_val in enumerate(gpa_cases):
        s = _make_student(310 + i)
        s["CUM_GPA"] = gpa_val
        c2_rows.append(s)
    _write_wb(c2_rows, out / "C2_malformed_gpa.xlsx")
    manifest.append(("C2_malformed_gpa.xlsx",
        "IMPORT — CUM_GPA values: valid, 0.0, 4.0, null, 'N/A', '3.5x', 5.0, -1.0, ''",
        f"new=11  (all rows imported; bad GPA stored as-is or null, logged as warning)",
        "Check normalization_warning events in import log for 'N/A', '3.5x', ''."))

    # ── C3: unicode / non-ASCII names ─────────────────────────────────────────
    c3_rows = _students(320, 5)
    c3_rows[0]["STUD_NAME"] = "أحمد علي"                  # Arabic
    c3_rows[0]["WSP_PREFERRED_TYPE_OF_WORK"] = "مساعد إداري"
    c3_rows[1]["STUD_NAME"] = "Élodie Châtaignier"        # French accented
    c3_rows[2]["STUD_NAME"] = "María-José Rodríguez"      # Spanish
    c3_rows[3]["STUD_NAME"] = "Ångström Björk"            # Nordic
    c3_rows[4]["STUD_NAME"] = "李 明"                      # Chinese
    c3_rows[4]["WSP_SPOKEN_LANGUAGES"] = "Arabic, English, Chinese (Mandarin)"
    _write_wb(c3_rows, out / "C3_unicode_names.xlsx")
    manifest.append(("C3_unicode_names.xlsx",
        "IMPORT — student names in Arabic, French, Spanish, Nordic, Chinese",
        "new=5  (all characters preserved)",
        "Check STUD_NAME in Excel Sheets — Arabic/accented/CJK must render correctly."))

    # ── C4: extremely long skill text fields ──────────────────────────────────
    long_skills = ("Python, SQL, FastAPI, React, Docker, Kubernetes, Terraform, AWS, " * 20).strip(", ")
    long_pref   = ("Software development, data engineering, backend systems, API design, " * 15).strip(", ")
    c4_rows = _students(330, 3)
    c4_rows[0]["WSP_TECHNICAL_SKILLS"] = long_skills        # ~1 800 chars
    c4_rows[1]["WSP_PREFERRED_TYPE_OF_WORK"] = long_pref    # ~1 100 chars
    c4_rows[2]["WSP_ORGANIZATIONAL_SKILLS"] = "A" * 2000   # 2 000 char blob
    _write_wb(c4_rows, out / "C4_long_text_fields.xlsx")
    manifest.append(("C4_long_text_fields.xlsx",
        "IMPORT — WSP skill fields with 1 000–2 000+ character values",
        "new=3  (long text stored; no truncation)",
        "Check that skills are fully stored in Excel Sheets → Column_Schema view."))

    # ── C5: whitespace-padded values / all-space values ──────────────────────
    c5_rows = _students(340, 5)
    c5_rows[0]["STUD_NAME"] = "  Padded Name  "
    c5_rows[1]["WSP_TECHNICAL_SKILLS"] = "   Excel   ,   SQL   "
    c5_rows[2]["WSP_ORGANIZATIONAL_SKILLS"] = "     "   # all spaces → should normalise to null
    c5_rows[3]["MAJR_DESC"] = "\tBusiness Administration\t"
    c5_rows[4]["MOBILE_NBR"] = "  +961 71 123456  "
    _write_wb(c5_rows, out / "C5_whitespace_values.xlsx")
    manifest.append(("C5_whitespace_values.xlsx",
        "IMPORT — values with leading/trailing spaces, tab padding, all-space fields",
        "new=5  (values trimmed; all-space fields stored as null or empty)",
        "Check STUD_NAME trimmed, all-space org skills normalised, tab stripped from major."))

    return manifest


# ══════════════════════════════════════════════════════════════════════════════
# GROUP D — Column schema variants
# ══════════════════════════════════════════════════════════════════════════════

def gen_D(out: Path) -> list[dict]:
    manifest = []

    # ── D1: extra unknown columns present ────────────────────────────────────
    extra_cols = COLS + ["DEPT_CODE", "SUPERVISOR_NAME", "CONTRACT_TYPE", "OFFICE_LOCATION"]
    d1_rows = _students(400, 8)
    for s in d1_rows:
        s["DEPT_CODE"] = f"DEPT{_RNG.randint(10,99)}"
        s["SUPERVISOR_NAME"] = f"Prof. {_RNG.choice(['Karam', 'Habib', 'Nassar'])}"
        s["CONTRACT_TYPE"] = _RNG.choice(["Part-time", "Full-time", "Volunteer"])
        s["OFFICE_LOCATION"] = _RNG.choice(["West Hall", "Bliss Hall", "College Hall"])
    _write_wb(d1_rows, out / "D1_extra_columns.xlsx", columns=extra_cols)
    manifest.append(("D1_extra_columns.xlsx",
        "IMPORT — all standard columns + 4 extra columns",
        "new=8  (extra cols stored in extra_columns_json, not rejected)",
        "Check Excel Sheets → Column_Schema tab shows new columns detected."))

    # ── D2: only STUD_ID + STUD_NAME + MAJR_DESC (minimal) ──────────────────
    minimal_cols = ["STUD_ID", "STUD_NAME", "MAJR_DESC"]
    d2_rows = [{c: _make_student(420 + i).get(c) for c in minimal_cols} for i in range(5)]
    _write_wb(d2_rows, out / "D2_minimal_columns.xlsx", columns=minimal_cols)
    manifest.append(("D2_minimal_columns.xlsx",
        "IMPORT — only 3 columns: STUD_ID, STUD_NAME, MAJR_DESC",
        "new=5  (all other fields null; schema warnings expected)",
        "Import succeeds. Missing-column warnings in log. WSP fields null in DB."))

    # ── D3: partial columns (~20 of 39) ──────────────────────────────────────
    partial_cols = ["STUD_ID", "STUD_NAME", "MAJR_DESC", "CLAS_DESC", "STUD_EMAIL",
                    "CUM_GPA", "PROBATION", "FINANCIAL_AID", "DORMS",
                    "WSP_TECHNICAL_SKILLS", "WSP_PREFERRED_TYPE_OF_WORK",
                    "ENRL_TERM", "MOBILE_NBR", "LEVL_CODE", "COLL_CODE",
                    "TOTAL_CREDIT_HOURS", "STST_DESC", "ENROLLED_IND",
                    "ASTD_DESC", "ATSD_CODE_END_OF_TERM"]
    d3_rows = [{c: _make_student(430 + i).get(c) for c in partial_cols} for i in range(8)]
    _write_wb(d3_rows, out / "D3_partial_columns.xlsx", columns=partial_cols)
    manifest.append(("D3_partial_columns.xlsx",
        "IMPORT — 20 of 39 columns (19 expected columns missing)",
        "new=8  (missing columns null; new_columns_detected=0; missing_columns_detected=19)",
        "Check System Status diagnostics. Schema comparison shows missing cols."))

    # ── D4: all headers lowercase ─────────────────────────────────────────────
    d4_rows = _students(450, 5)
    lower_cols = [c.lower() for c in COLS]
    _write_wb(d4_rows, out / "D4_lowercase_headers.xlsx", columns=lower_cols, data_keys=COLS)
    manifest.append(("D4_lowercase_headers.xlsx",
        "IMPORT — all column headers in lowercase (stud_id, stud_name, …)",
        "new=5  (normalize_header() uppercases and strips; import succeeds)",
        "Students imported correctly despite lowercase headers."))

    # ── D5: mixed-case / spaced headers ──────────────────────────────────────
    mixed_headers = [
        "Stud Id", "Stud Name", "Majr Desc", "Clas Desc", "Stud Email",
        "WSP Written Languages", "WSP Spoken Languages",
        "WSP_Organizational_Skills", "WSP Technical Skills",
        "WSP_Interpersonal_Skills", "WSP Additional Skills",
        "WSP Prev Work", "WSP Previous Type Of Work", "WSP Preferred Type Of Work",
        "Deans Warning", "Enrl Term", "Dean Warn", "Mobile Nbr", "Probation",
        "Application Date", "Cum GPA", "STST Desc", "STYP Code", "STYP Desc",
        "Enrolled Ind", "Registered Ind", "LEVL Code", "COLL Code",
        "Total Credit Hours", "ASTD Term", "Atsd Code End Of Term", "ASTD Desc",
        "ASTD Date End Of Term", "USAID", "Master Card", "UPP MEPI", "GAS",
        "Financial Aid", "DORMS",
    ]
    d5_rows = _students(460, 5)
    _write_wb(d5_rows, out / "D5_mixed_case_headers.xlsx", columns=mixed_headers, data_keys=COLS)
    manifest.append(("D5_mixed_case_headers.xlsx",
        "IMPORT — headers in mixed case with spaces (normalize_header maps to standard names)",
        "new=5  (header normalisation maps e.g. 'Cum GPA' → 'CUM_GPA')",
        "Students imported. Confirm correct field mapping via Excel Sheets view."))

    return manifest


# ══════════════════════════════════════════════════════════════════════════════
# GROUP E — Sheet / workbook structure variants
# ══════════════════════════════════════════════════════════════════════════════

def gen_E(out: Path) -> list[dict]:
    manifest = []

    # ── E1: multi-sheet workbook (students on Sheet1 = first) ────────────────
    e1_students = _students(500, 15)
    wb = Workbook()
    ws1 = wb.active; ws1.title = "Sheet1"
    ws1.append(COLS)
    for s in e1_students: ws1.append([s.get(c) for c in COLS])
    ws2 = wb.create_sheet("Major_Summary")
    ws2.append(["Major", "Count"])
    ws2.append(["Computer Science", 5])
    ws3 = wb.create_sheet("Aid_Roster")
    ws3.append(["STUD_ID", "STUD_NAME", "FINANCIAL_AID"])
    for s in e1_students[:3]: ws3.append([s["STUD_ID"], s["STUD_NAME"], "Y"])
    wb.save(out / "E1_multi_sheet_students_first.xlsx")
    manifest.append(("E1_multi_sheet_students_first.xlsx",
        "IMPORT — 3 sheets: Sheet1 (students), Major_Summary, Aid_Roster",
        "new=15  (importer reads Sheet1 = first sheet only)",
        "Sheet names shown in workbook picker. Only Sheet1 data imported."))

    # ── E2: multi-sheet, student data NOT on first sheet ─────────────────────
    wb2 = Workbook()
    ws_info = wb2.active; ws_info.title = "Info"
    ws_info["A1"] = "This workbook has student data on Sheet2, not Sheet1."
    ws_std = wb2.create_sheet("Students")
    ws_std.append(COLS)
    for s in _students(520, 5): ws_std.append([s.get(c) for c in COLS])
    wb2.save(out / "E2_students_not_on_first_sheet.xlsx")
    manifest.append(("E2_students_not_on_first_sheet.xlsx",
        "IMPORT — student data on 2nd sheet ('Students'), not 1st ('Info')",
        "Expected: 0 students imported (importer reads first sheet 'Info' which has no STUD_ID)",
        "Import may warn: NoValidStudentRowsError or 0 new rows. Verify import log."))

    # ── E3: headers only — zero data rows ────────────────────────────────────
    _write_wb([], out / "E3_empty_no_rows.xlsx")
    manifest.append(("E3_empty_no_rows.xlsx",
        "IMPORT — header row present, ZERO data rows  ← should fail gracefully",
        "Expected: NoValidStudentRowsError — import rejected, nothing written",
        "Import page shows error. DB unchanged."))

    # ── E4: single student ────────────────────────────────────────────────────
    _write_wb(_students(600, 1), out / "E4_single_student.xlsx")
    manifest.append(("E4_single_student.xlsx",
        "IMPORT — exactly 1 student",
        "new=1  (minimal-size import works)",
        "Dashboard shows 1 student (or existing + 1 if prior imports)."))

    # ── E5: blank rows interspersed in data ──────────────────────────────────
    e5_real = _students(610, 8)
    wb5 = Workbook()
    ws5 = wb5.active; ws5.title = "Sheet1"
    ws5.append(COLS)
    for i, s in enumerate(e5_real):
        ws5.append([s.get(c) for c in COLS])
        if i in (2, 5):                     # insert blank rows after rows 2 and 5
            ws5.append([None] * len(COLS))
    wb5.save(out / "E5_blank_rows_interspersed.xlsx")
    manifest.append(("E5_blank_rows_interspersed.xlsx",
        "IMPORT — 8 valid data rows with 2 blank rows interspersed",
        "new=8  (blank rows skipped; only rows with STUD_ID processed)",
        "Exactly 8 students imported, 2 blank rows logged as skipped (no STUD_ID)."))

    # ── E6: 100-student medium-size import ────────────────────────────────────
    _write_wb(_students(700, 100), out / "E6_medium_100_students.xlsx")
    manifest.append(("E6_medium_100_students.xlsx",
        "IMPORT — 100 students (medium-size, all fields populated)",
        "new=100  (standard import, performance check)",
        "Dashboard updates to 100+ total. Import should complete in < 5 seconds."))

    # ── E7: 500-student large import ─────────────────────────────────────────
    _write_wb(_students(800, 500), out / "E7_large_500_students.xlsx")
    manifest.append(("E7_large_500_students.xlsx",
        "IMPORT — 500 students (large import, stress test)",
        "new=500  (all imported; check UI stays responsive)",
        "Dashboard shows 500+ students. Import log has 500 entries."))

    return manifest


# ══════════════════════════════════════════════════════════════════════════════
# GROUP F — Filter & semantic search test targets
# ══════════════════════════════════════════════════════════════════════════════

def gen_F(out: Path) -> list[dict]:
    manifest = []

    f_rows: list[dict] = []
    sid = 900

    def _add(label, **kw) -> dict:
        nonlocal sid
        s = _make_student(sid, **kw)
        s["_label"] = label       # cosmetic only — won't be in COLS
        f_rows.append(s)
        sid += 1
        return s

    # ── probation cohort (10) ─────────────────────────────────────────────────
    for _ in range(10):
        _add("PROBATION", prob=True, aid=False, dorms=False)

    # ── financial aid only, no probation (10) ────────────────────────────────
    for _ in range(10):
        _add("AID_ONLY", prob=False, aid=True, dorms=False)

    # ── dorms only (10) ──────────────────────────────────────────────────────
    for _ in range(10):
        _add("DORMS", prob=False, aid=False, dorms=True)

    # ── aid + dorms combo (5) ────────────────────────────────────────────────
    for _ in range(5):
        _add("AID+DORMS", prob=False, aid=True, dorms=True)

    # ── one per college (8 major groups) ─────────────────────────────────────
    for i, (majr, coll, skills, pref) in enumerate(_MAJORS[:8]):
        s = _make_student(sid)
        s["MAJR_DESC"] = majr; s["COLL_CODE"] = coll
        s["WSP_TECHNICAL_SKILLS"] = skills
        s["WSP_PREFERRED_TYPE_OF_WORK"] = pref
        f_rows.append(s); sid += 1

    # ── one per class year ───────────────────────────────────────────────────
    for cls in _CLASSES:
        s = _make_student(sid, cls=cls)
        f_rows.append(s); sid += 1

    # ── GPA ranges for filter testing ────────────────────────────────────────
    gpa_brackets = [(1.5, 1.99), (2.0, 2.49), (2.5, 2.99), (3.0, 3.49), (3.5, 4.0)]
    for lo, hi in gpa_brackets:
        s = _make_student(sid)
        s["CUM_GPA"] = round(_RNG.uniform(lo, hi), 2)
        f_rows.append(s); sid += 1

    # ── semantic search targets ───────────────────────────────────────────────
    # Strong match for "data entry spreadsheet reporting"
    s1 = _make_student(sid)
    s1["STUD_NAME"] = "Maya Spreadsheet"
    s1["WSP_TECHNICAL_SKILLS"] = "Excel, SQL, data entry, spreadsheet QA, VLOOKUP, pivot tables"
    s1["WSP_PREFERRED_TYPE_OF_WORK"] = "Data entry, spreadsheet reporting, import error checking"
    s1["WSP_ORGANIZATIONAL_SKILLS"] = "Record keeping, data validation, documentation"
    f_rows.append(s1); sid += 1

    # Strong match for "laboratory research biology"
    s2 = _make_student(sid)
    s2["STUD_NAME"] = "Omar LabWork"
    s2["MAJR_DESC"] = "Biology"
    s2["WSP_TECHNICAL_SKILLS"] = "PCR, gel electrophoresis, cell culture, CRISPR, ELISA"
    s2["WSP_PREFERRED_TYPE_OF_WORK"] = "Laboratory research, biological experiments, sample analysis"
    f_rows.append(s2); sid += 1

    # Strong match for "graphic design social media content"
    s3 = _make_student(sid)
    s3["STUD_NAME"] = "Lara Design"
    s3["MAJR_DESC"] = "Graphic Design"
    s3["WSP_TECHNICAL_SKILLS"] = "Adobe Illustrator, Photoshop, Figma, Canva, After Effects"
    s3["WSP_PREFERRED_TYPE_OF_WORK"] = "Graphic design, social media content creation, branding"
    f_rows.append(s3); sid += 1

    # Strong match for "nursing clinical patient care"
    s4 = _make_student(sid)
    s4["STUD_NAME"] = "Hana ClinicalCare"
    s4["MAJR_DESC"] = "Nursing"
    s4["WSP_TECHNICAL_SKILLS"] = "IV insertion, patient assessment, vital signs, wound care, EMR"
    s4["WSP_PREFERRED_TYPE_OF_WORK"] = "Clinical support, patient care, hospital ward assistance"
    f_rows.append(s4); sid += 1

    # Strong match for "software engineering Python backend"
    s5 = _make_student(sid)
    s5["STUD_NAME"] = "Jad Backend"
    s5["MAJR_DESC"] = "Computer Science"
    s5["WSP_TECHNICAL_SKILLS"] = "Python, FastAPI, PostgreSQL, Docker, Git, REST APIs, unit testing"
    s5["WSP_PREFERRED_TYPE_OF_WORK"] = "Backend software development, API design, code review"
    f_rows.append(s5); sid += 1

    # ── known-name targets for name search ────────────────────────────────────
    for name in ["Zebra Unique", "Unicorn Test", "Delta Filter", "Alpha Search"]:
        s = _make_student(sid)
        s["STUD_NAME"] = name
        f_rows.append(s); sid += 1

    # strip internal _label key before writing
    clean_rows = [{c: s.get(c) for c in COLS} for s in f_rows]
    _write_wb(clean_rows, out / "F1_filter_and_search_targets.xlsx")

    manifest.append(("F1_filter_and_search_targets.xlsx",
        "IMPORT — 72 students designed to test every filter + semantic search",
        "new=72  (all types: probation, aid, dorms, GPA ranges, class years, colleges)",
        "\n".join([
            "  Filter tests to run after importing:",
            "  • PROBATION=Y         → expect 10 results",
            "  • FINANCIAL_AID=Y     → expect 15 results (aid-only + aid+dorms)",
            "  • DORMS=Y             → expect 15 results (dorms-only + aid+dorms)",
            "  • GPA 1.5–2.0         → expect 1 result",
            "  • GPA 3.5–4.0         → expect 1+ result",
            "  • Class=Freshman      → expect 1 result",
            "  • Major=Biology       → expect 1+ result (s2 + any procedural Biology students)",
            "  • Name contains 'Zebra' → expect exactly 1 result",
            "  • Name contains 'Unique' → expect exactly 1 result",
            "  Semantic search tests:",
            "  • 'data entry spreadsheet reporting' → Maya Spreadsheet should rank #1",
            "  • 'laboratory research biology'      → Omar LabWork should rank #1",
            "  • 'graphic design social media'      → Lara Design should rank #1",
            "  • 'nursing clinical patient care'    → Hana ClinicalCare should rank #1",
            "  • 'software engineering Python'      → Jad Backend should rank #1",
        ])))

    return manifest


# ══════════════════════════════════════════════════════════════════════════════
# GROUP G — Cumulative waves to 3 000 active students
# ══════════════════════════════════════════════════════════════════════════════

def gen_G(out: Path) -> list[tuple]:
    manifest = []

    # Each wave carries all previous students forward and adds a new cohort.
    # IDs start at 5000 so they never collide with groups A-F (max ~1400).
    WAVE_SIZES   = [500, 500, 500, 500, 500, 500]   # new students per wave
    cumulative:  list[dict] = []
    next_id      = 5000
    running_total = 0

    for wave_idx, cohort_size in enumerate(WAVE_SIZES, start=1):
        wave_letter = chr(ord("A") + wave_idx - 1)   # A … F

        # Update 10 % of carry-over students each wave (skip wave 1)
        updated_ids = set()
        if wave_idx > 1:
            n_update = max(1, len(cumulative) // 10)
            for s in _RNG.sample(cumulative, n_update):
                s["CUM_GPA"]              = round(min(4.0, s["CUM_GPA"] + _RNG.uniform(-0.15, 0.20)), 2)
                s["WSP_TECHNICAL_SKILLS"] = s["WSP_TECHNICAL_SKILLS"].rstrip(", ") + f", updated-w{wave_idx}"
                updated_ids.add(s["STUD_ID"])

        # New cohort — realistic distribution:
        #   ~15% on academic probation  (GPA 1.50–2.29, PROBATION=Y)
        #   ~25% on financial aid       (FINANCIAL_AID=Y, not necessarily probation)
        #   ~20% in dorms               (DORMS=Y, independent of above)
        new_cohort = []
        for i in range(cohort_size):
            roll = _RNG.random()
            prob  = roll < 0.15
            aid   = _RNG.random() < 0.25
            dorms = _RNG.random() < 0.20
            new_cohort.append(_make_student(next_id + i, prob=prob, aid=aid, dorms=dorms))
        next_id += cohort_size

        cumulative += new_cohort
        running_total = len(cumulative)

        fname = f"G{wave_idx}_wave{wave_letter}_{running_total}_students.xlsx"
        _write_wb(cumulative, out / fname)

        updated_str = f"updated={len(updated_ids)}" if updated_ids else "updated=0"
        manifest.append((fname,
            f"IMPORT WAVE {wave_idx} — cumulative total {running_total} active students",
            f"new={cohort_size}  {updated_str}  unchanged={running_total - cohort_size - len(updated_ids)}",
            f"Dashboard active count should reach {running_total}. Import log shows {cohort_size} new rows."))

    return manifest


# ══════════════════════════════════════════════════════════════════════════════
# MAIN
# ══════════════════════════════════════════════════════════════════════════════

def main() -> None:
    import sys
    if hasattr(sys.stdout, "reconfigure"):
        sys.stdout.reconfigure(encoding="utf-8", errors="replace")

    out = Path(__file__).resolve().parent.parent / "data" / "testbench"
    if out.exists():
        shutil.rmtree(out)
    out.mkdir(parents=True)
    print(f"Writing testbench workbooks to: {out}\n")

    manifest: list[tuple] = []
    manifest += gen_A(out)
    manifest += gen_B(out)
    manifest += gen_C(out)
    manifest += gen_D(out)
    manifest += gen_E(out)
    manifest += gen_F(out)
    manifest += gen_G(out)

    # ── print manifest ────────────────────────────────────────────────────────
    files = sorted(out.glob("*.xlsx"))
    SEP = "-" * 72
    print(SEP)
    print(f"Generated {len(files)} workbooks\n")

    for filename, action, expected, verify in manifest:
        size_kb = (out / filename).stat().st_size // 1024
        print(f"  [{filename}]  ({size_kb} KB)")
        print(f"    Action  : {action}")
        print(f"    Expected: {expected}")
        print(f"    Verify  : {verify}")
        print()

    print(SEP)
    print("RECOMMENDED IMPORT ORDER FOR FULL COVERAGE")
    print(SEP)
    order = [
        ("1",  "A1  — clean first import, 30 students"),
        ("2",  "A2  — add 10 + update 5"),
        ("3",  "A3  — 8 go missing + 2 new"),
        ("4",  "A4  — restore 8 missing (same data → restored_student)"),
        ("5",  "A5  — restore again with changes (→ updated_student)"),
        ("6",  "A6  — try importing A1 again → DUPLICATE REJECTION"),
        ("7",  "B1  — integer STUD_IDs"),
        ("8",  "B2  — float STUD_IDs"),
        ("9",  "B3  — leading-zero IDs"),
        ("10", "B4  — rows with missing STUD_ID"),
        ("11", "C1  — boolean variants (Y/N vs TRUE/FALSE vs 1/0)"),
        ("12", "C2  — malformed GPA values"),
        ("13", "C3  — unicode/Arabic/French names"),
        ("14", "C4  — very long skill text"),
        ("15", "C5  — whitespace-padded values"),
        ("16", "D1  — extra columns → extra_columns_json"),
        ("17", "D2  — minimal 3-column workbook"),
        ("18", "D3  — partial columns (20 of 39)"),
        ("19", "D4  — all lowercase headers"),
        ("20", "D5  — mixed-case headers with spaces"),
        ("21", "E1  — multi-sheet (students on first sheet)"),
        ("22", "E2  — multi-sheet (students NOT on first sheet)"),
        ("23", "E3  — empty workbook (headers only) → REJECTION"),
        ("24", "E4  — single student"),
        ("25", "E5  — blank rows interspersed"),
        ("26", "E6  — 100 students (medium)"),
        ("27", "E7  — 500 students (large/stress)"),
        ("28", "F1  — filter & semantic search targets"),
        ("29", "G1  — wave A: 500 new students (cumulative: 500)"),
        ("30", "G2  — wave B: 500 new + 10% updates (cumulative: 1 000)"),
        ("31", "G3  — wave C: 500 new + 10% updates (cumulative: 1 500)"),
        ("32", "G4  — wave D: 500 new + 10% updates (cumulative: 2 000)"),
        ("33", "G5  — wave E: 500 new + 10% updates (cumulative: 2 500)"),
        ("34", "G6  — wave F: 500 new + 10% updates (cumulative: 3 000)  <-- 3k milestone"),
    ]
    for n, desc in order:
        print(f"  {n:>2}. {desc}")
    print()


if __name__ == "__main__":
    main()
